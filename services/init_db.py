# services/init_db.py

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime
from flask_login import UserMixin
from config import DATABASE_URL, UPLOAD_FOLDER, ES_HOST, ES_USER, ES_PASS, INVENTORY_IMAGE_FOLDER
from elasticsearch import Elasticsearch
from werkzeug.security import generate_password_hash
import logging
import json
import openpyxl
from services.db import get_connection, release_connection, db_connection, db_transaction

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('init_db.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# Add the parent directory to Python path to import config
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

es = Elasticsearch([ES_HOST], basic_auth=(ES_USER, ES_PASS))

# Initialize Elasticsearch index
if not es.indices.exists(index='knowledge_base'):
    es.indices.create(index='knowledge_base', body={
        'mappings': {
            'properties': {
                'text': {'type': 'text'},
                'filename': {'type': 'keyword'}
            }
        }
    })

def get_database_connection():
    """Get database connection with error handling"""
    try:
        conn = get_connection()
        logger.info("✅ Database connection successful")
        return conn
    except psycopg2.OperationalError as e:
        logger.error(f"❌ Database connection failed: {e}")
        logger.error("Check your DATABASE_URL in config.py or environment variables")
        raise
    except Exception as e:
        logger.error(f"❌ Unexpected database error: {e}")
        raise

def upgrade_user_table_for_enterprise():
    """Upgrade users table for enterprise features"""
    conn = get_database_connection()
    c = conn.cursor()
    
    try:
        logger.info("🔧 Upgrading users table...")
        
        # Check existing columns
        c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'users'")
        existing_columns = [row[0] for row in c.fetchall()]
        
        # Columns to add
        new_columns = [
            ("full_name", "TEXT"),
            ("job_title", "TEXT"),
            ("role", "TEXT DEFAULT 'user'"),
            ("certifications", "TEXT DEFAULT '[]'"),
            ("password", "TEXT"),
            ("timestamp", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        ]
        
        for col_name, col_definition in new_columns:
            if col_name not in existing_columns:
                c.execute(f"ALTER TABLE users ADD COLUMN {col_name} {col_definition}")
                logger.info(f"  ✅ Added column: {col_name}")
        
        conn.commit()
        logger.info("✅ Users table upgraded successfully")
        
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error upgrading users table: {e}")
        raise
    finally:
        release_connection(conn)

def upgrade_equipment_instances_table():
    """Upgrade equipment instances table"""
    conn = get_database_connection()
    c = conn.cursor()
    
    try:
        logger.info("🔧 Upgrading equipment_instances table...")
        
        c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'equipment_instances'")
        existing_columns = [row[0] for row in c.fetchall()]
        
        columns_to_add = [
            ("equipment_type", "TEXT"),
            ("unique_id", "TEXT"),
            ("brand", "TEXT"),
            ("model", "TEXT"),
            ("serial_number", "TEXT"),
            ("hours", "REAL DEFAULT 0"),
            ("fuel_type", "TEXT"),
            ("gross_weight", "REAL"),
            ("requires_operator", "BOOLEAN DEFAULT FALSE"),
            ("required_certification", "TEXT"),
            ("status", "TEXT DEFAULT 'available'"),
            ("last_maintenance", "TIMESTAMP"),
            ("maintenance_threshold", "INTEGER")
        ]
        
        # Check if table has rows
        c.execute("SELECT COUNT(*) FROM equipment_instances")
        row_count = c.fetchone()[0]
        
        for col_name, col_type in columns_to_add:
            if col_name not in existing_columns:
                c.execute(f"ALTER TABLE equipment_instances ADD COLUMN {col_name} {col_type}")
                logger.info(f"  ✅ Added column: {col_name}")
        
        # Update existing rows with default values for NOT NULL columns
        if row_count > 0:
            c.execute("UPDATE equipment_instances SET equipment_type = 'Unknown' WHERE equipment_type IS NULL")
            c.execute("UPDATE equipment_instances SET unique_id = COALESCE(unique_id, id::TEXT) WHERE unique_id IS NULL")
        
        # Add NOT NULL constraints (carefully)
        if 'equipment_type' in [col[0] for col in columns_to_add]:
            c.execute("ALTER TABLE equipment_instances ALTER COLUMN equipment_type SET NOT NULL")
        if 'unique_id' in [col[0] for col in columns_to_add]:
            c.execute("ALTER TABLE equipment_instances ALTER COLUMN unique_id SET NOT NULL")
        
        # Add UNIQUE constraint on unique_id if not exists
        try:
            c.execute("ALTER TABLE equipment_instances ADD CONSTRAINT unique_id_unique UNIQUE (unique_id)")
        except psycopg2.errors.DuplicateTable:
            pass  # Constraint already exists
        
        conn.commit()
        logger.info("✅ Equipment instances table upgraded successfully")
        
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error upgrading equipment_instances table: {e}")
        raise
    finally:
        release_connection(conn)

def create_billing_tables():
    """Create tables for service units and invoices."""
    conn = None
    try:
        conn = get_connection()
        conn.autocommit = True
        c = conn.cursor()

        # Service Units from Excel
        c.execute('''
            CREATE TABLE IF NOT EXISTS service_units (
                id SERIAL PRIMARY KEY,
                source_file TEXT,
                service_unit TEXT NOT NULL,
                core_code TEXT,
                service_master TEXT,
                unit_of_measure TEXT,
                definition TEXT,
                clarification TEXT,
                sap_service_master_ TEXT,
                task_code_description TEXT,
                unnamed_2 TEXT,
                description TEXT,
                uom TEXT,
                explanation TEXT,
                om_contract_service_units TEXT,
                core_service_master TEXT,
                non_core_code TEXT,
                non_core_service_master TEXT,
                work_type TEXT,
                definitions TEXT,
                category TEXT,  -- e.g., 'NCC', 'System', etc.
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Invoices
        c.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                job_id INTEGER REFERENCES jobs(id) ON DELETE SET NULL,
                invoice_number TEXT UNIQUE,
                title TEXT,
                customer_id INTEGER REFERENCES maintenance_customers(id),
                status TEXT DEFAULT 'draft',
                subtotal NUMERIC(12,2) DEFAULT 0,
                tax NUMERIC(12,2) DEFAULT 0,
                total NUMERIC(12,2) DEFAULT 0,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                sent_at TIMESTAMP,
                paid_at TIMESTAMP
            )
        ''')

        # Invoice line items (linked to service units)
        c.execute('''
            CREATE TABLE IF NOT EXISTS invoice_lines (
                id SERIAL PRIMARY KEY,
                invoice_id INTEGER REFERENCES invoices(id) ON DELETE CASCADE,
                service_unit_id INTEGER REFERENCES service_units(id),
                description TEXT,
                quantity NUMERIC(10,2) DEFAULT 1,
                unit_price NUMERIC(12,2) DEFAULT 0,
                total NUMERIC(12,2) DEFAULT 0,
                position INTEGER DEFAULT 0
            )
        ''')

    except Exception as e:
        logging.error(f"Error creating billing tables: {str(e)}")
    finally:
        try:
            if conn:
                release_connection(conn)
        except Exception:
            pass

def upgrade_consumables_table():
    """Enhanced consumables table with all new inventory features"""
    conn = get_database_connection()
    c = conn.cursor()
    
    try:
        logger.info("🔧 Upgrading consumables table with inventory enhancements...")
        
        # Check existing columns
        c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'consumables'")
        existing_columns = [row[0] for row in c.fetchall()]
        
        # Add new columns for enhanced inventory management
        new_columns = [
            ("category", "TEXT"),
            ("base_name", "TEXT"),
            ("specifications", "JSONB"),
            ("manufacturer", "TEXT"),
            ("part_number", "TEXT"),
            ("supplier_sku", "TEXT"),
            ("internal_sku", "TEXT"),
            ("bin_location", "TEXT"),
            ("cost_per_unit", "DECIMAL(10,4)"),
            ("weight_per_unit", "DECIMAL(10,4)"),
            ("dimensions", "JSONB"),
            ("barcode", "TEXT"),
            ("image_path", "TEXT"),
            ("image_filename", "TEXT"),
            ("minimum_stock_level", "INTEGER"),
            ("maximum_stock_level", "INTEGER"),
            ("hazmat", "BOOLEAN DEFAULT FALSE"),
            ("requires_certification", "BOOLEAN DEFAULT FALSE"),
            ("product_url", "TEXT"),
            ("notes", "TEXT"),
            ("normalized_name", "TEXT"),
            ("active", "BOOLEAN DEFAULT TRUE"),
            ("created_at", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
            ("updated_at", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        ]
        
        added_count = 0
        for col_name, col_definition in new_columns:
            if col_name not in existing_columns:
                try:
                    c.execute(f"ALTER TABLE consumables ADD COLUMN {col_name} {col_definition}")
                    added_count += 1
                    logger.info(f"  ✅ Added column: {col_name}")
                except Exception as e:
                    logger.warning(f"  ⚠️  Could not add column {col_name}: {str(e)}")
        
        logger.info(f"✅ Added {added_count} new columns to consumables table")
        
        # Update existing rows with defaults
        c.execute("SELECT COUNT(*) FROM consumables")
        row_count = c.fetchone()[0]
        
        if row_count > 0:
            c.execute("UPDATE consumables SET category = 'general' WHERE category IS NULL")
            c.execute("UPDATE consumables SET base_name = name WHERE base_name IS NULL AND name IS NOT NULL")
            c.execute("UPDATE consumables SET normalized_name = LOWER(TRIM(name)) WHERE normalized_name IS NULL AND name IS NOT NULL")
            c.execute("UPDATE consumables SET specifications = '{}' WHERE specifications IS NULL")
            c.execute("UPDATE consumables SET active = TRUE WHERE active IS NULL")
            logger.info("  ✅ Updated existing rows with default values")
        
        # Create indexes for performance
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_consumables_category ON consumables(category)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_base_name ON consumables(base_name)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_supplier ON consumables(supplier)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_internal_sku ON consumables(internal_sku)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_supplier_sku ON consumables(supplier_sku)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_active ON consumables(active)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_specifications ON consumables USING GIN(specifications)"
        ]
        
        for index_sql in indexes:
            try:
                c.execute(index_sql)
            except Exception as e:
                logger.warning(f"Could not create index: {str(e)}")
        
        logger.info("✅ Consumables table indexes created")
        
        conn.commit()
        logger.info("✅ Consumables table upgrade completed successfully")
        
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error upgrading consumables table: {e}")
        raise
    finally:
        release_connection(conn)

def create_item_templates_table():
    """Create table for standardized item templates"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS item_templates (
        id SERIAL PRIMARY KEY,
        category TEXT NOT NULL,
        base_name TEXT NOT NULL,
        required_specs JSONB NOT NULL, -- Required specification fields
        optional_specs JSONB, -- Optional specification fields
        default_unit TEXT,
        naming_pattern TEXT NOT NULL, -- Pattern for generating standardized names
        description TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(category, base_name)
    )''')
    
    # Insert standard templates
    templates = [
        # Bolts template
        {
            'category': 'fasteners',
            'base_name': 'bolt',
            'required_specs': {
                'diameter': {'type': 'string', 'values': ['1/4"', '3/8"', '1/2"', '5/8"', '3/4"', '1"']},
                'length': {'type': 'string', 'pattern': r'^\d+(\.\d+)?"$'},
                'material': {'type': 'string', 'values': ['steel', 'stainless_steel', 'galvanized', 'brass']},
                'thread_type': {'type': 'string', 'values': ['coarse', 'fine']}
            },
            'optional_specs': {
                'head_type': {'type': 'string', 'values': ['hex', 'carriage', 'socket']},
                'coating': {'type': 'string', 'values': ['zinc', 'chrome', 'none']}
            },
            'default_unit': 'units',
            'naming_pattern': '{diameter} x {length} {material} {thread_type} bolt',
            'description': 'Standard bolts with diameter, length, material, and thread specifications'
        },
        # Cable template
        {
            'category': 'electrical',
            'base_name': 'cable',
            'required_specs': {
                'diameter': {'type': 'string', 'pattern': r'^\d+(\.\d+)?(mm|in)$'},
                'conductor_material': {'type': 'string', 'values': ['copper', 'aluminum', 'steel']},
                'construction': {'type': 'string', 'values': ['solid', 'stranded', 'braided']},
                'insulation': {'type': 'string', 'values': ['pvc', 'xlpe', 'rubber', 'bare']}
            },
            'optional_specs': {
                'voltage_rating': {'type': 'string', 'pattern': r'^\d+kV$'},
                'conductor_count': {'type': 'integer', 'min': 1, 'max': 50}
            },
            'default_unit': 'feet',
            'naming_pattern': '{diameter} {conductor_material} {construction} {insulation} cable',
            'description': 'Electrical cables with diameter, material, construction, and insulation specs'
        },
        # Safety equipment template
        {
            'category': 'safety',
            'base_name': 'gloves',
            'required_specs': {
                'material': {'type': 'string', 'values': ['rubber', 'leather', 'nitrile', 'latex']},
                'protection_class': {'type': 'string', 'values': ['class_0', 'class_1', 'class_2', 'class_3', 'class_4']},
                'size': {'type': 'string', 'values': ['XS', 'S', 'M', 'L', 'XL', 'XXL']}
            },
            'optional_specs': {
                'cuff_length': {'type': 'string', 'values': ['short', 'medium', 'long']},
                'color': {'type': 'string'}
            },
            'default_unit': 'pairs',
            'naming_pattern': '{material} {protection_class} {size} safety gloves',
            'description': 'Safety gloves with material, protection class, and size specifications'
        }
    ]
    
    for template in templates:
        c.execute("""
            INSERT INTO item_templates 
            (category, base_name, required_specs, optional_specs, default_unit, naming_pattern, description)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (category, base_name) DO UPDATE SET
                required_specs = EXCLUDED.required_specs,
                optional_specs = EXCLUDED.optional_specs,
                default_unit = EXCLUDED.default_unit,
                naming_pattern = EXCLUDED.naming_pattern,
                description = EXCLUDED.description
        """, (
            template['category'],
            template['base_name'],
            json.dumps(template['required_specs']),
            json.dumps(template.get('optional_specs', {})),
            template['default_unit'],
            template['naming_pattern'],
            template['description']
        ))
    
    conn.commit()
    release_connection(conn)

def create_supplier_items_table():
    """Create table to track supplier-specific variations"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS supplier_items (
        id SERIAL PRIMARY KEY,
        consumable_id INTEGER NOT NULL,
        supplier_name TEXT NOT NULL,
        supplier_part_number TEXT,
        supplier_price DECIMAL(10,2),
        lead_time_days INTEGER,
        minimum_order_quantity INTEGER DEFAULT 1,
        preferred BOOLEAN DEFAULT FALSE,
        notes TEXT,
        last_ordered TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
        UNIQUE(consumable_id, supplier_name)
    )''')
    
    # Create indexes
    c.execute("CREATE INDEX IF NOT EXISTS idx_supplier_items_consumable ON supplier_items(consumable_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_supplier_items_supplier ON supplier_items(supplier_name)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_supplier_items_preferred ON supplier_items(preferred)")
    
    conn.commit()
    release_connection(conn)

def upgrade_integrations_table():
    """Create or upgrade integrations table"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS integrations (
        id SERIAL PRIMARY KEY,
        user_id INTEGER NOT NULL,
        system_type TEXT NOT NULL,
        system_name TEXT,
        config JSONB,
        enabled BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    
    c.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_user_system_integration 
        ON integrations(user_id, system_type)
    ''')
    
    conn.commit()
    release_connection(conn)

def create_reorder_tables():
    """Create tables to support reorder suggest→approve→dispatch workflow"""
    conn = get_connection()
    c = conn.cursor()

    # Suggested reorders generated by system or user
    c.execute('''CREATE TABLE IF NOT EXISTS reorder_requests (
        id SERIAL PRIMARY KEY,
        consumable_id INTEGER NOT NULL,
        suggested_quantity INTEGER NOT NULL,
        unit TEXT,
        reason TEXT,
        preferred_supplier_id INTEGER,
        estimated_cost DECIMAL(12,2),
        lead_time_days INTEGER,
        status TEXT DEFAULT 'pending',
        created_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        idempotency_key TEXT,
        UNIQUE(idempotency_key),
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE
    )''')

    # Approval records
    c.execute('''CREATE TABLE IF NOT EXISTS reorder_approvals (
        id SERIAL PRIMARY KEY,
        request_id INTEGER NOT NULL,
        approved_by INTEGER NOT NULL,
        approved_quantity INTEGER NOT NULL,
        supplier_id INTEGER,
        approval_notes TEXT,
        status TEXT DEFAULT 'approved',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (request_id) REFERENCES reorder_requests(id) ON DELETE CASCADE
    )''')

    # Ensure purchase_orders exists and has columns to link to reorder_requests
    # 1) Create table if not exists with a broad schema
    c.execute('''CREATE TABLE IF NOT EXISTS purchase_orders (
        id SERIAL PRIMARY KEY,
        po_number TEXT,
        supplier_id INTEGER,
        status TEXT DEFAULT 'draft',
        order_date DATE,
        expected_delivery_date DATE,
        actual_delivery_date DATE,
        subtotal DECIMAL(10,2),
        tax_amount DECIMAL(10,2),
        shipping_cost DECIMAL(10,2),
        total_amount DECIMAL(10,2),
        notes TEXT,
        created_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    # 2) Add missing columns used by reorder workflow if they don't exist
    c.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE table_name = 'purchase_orders'
    """)
    existing_po_cols = {row[0] for row in c.fetchall()}

    alter_po_cols = []
    if 'request_id' not in existing_po_cols:
        alter_po_cols.append("ADD COLUMN request_id INTEGER")
    if 'supplier_name' not in existing_po_cols:
        alter_po_cols.append("ADD COLUMN supplier_name TEXT")
    if 'payload' not in existing_po_cols:
        alter_po_cols.append("ADD COLUMN payload JSONB")
    if 'dispatch_channel' not in existing_po_cols:
        alter_po_cols.append("ADD COLUMN dispatch_channel TEXT")
    if 'error' not in existing_po_cols:
        alter_po_cols.append("ADD COLUMN error TEXT")

    if alter_po_cols:
        c.execute(f"ALTER TABLE purchase_orders {', '.join(alter_po_cols)}")
        # Add FK if request_id added
        if 'request_id' not in existing_po_cols:
            try:
                c.execute("ALTER TABLE purchase_orders ADD CONSTRAINT fk_po_request FOREIGN KEY (request_id) REFERENCES reorder_requests(id) ON DELETE SET NULL")
            except Exception:
                pass

    # Helpful indexes
    c.execute("CREATE INDEX IF NOT EXISTS idx_reorder_requests_status ON reorder_requests(status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_reorder_requests_consumable ON reorder_requests(consumable_id)")

    # Create index on purchase_orders.request_id only if column exists
    if 'request_id' in existing_po_cols or 'request_id' in {col.split()[2] for col in alter_po_cols}:
        try:
            c.execute("CREATE INDEX IF NOT EXISTS idx_purchase_orders_request ON purchase_orders(request_id)")
        except Exception:
            pass

    # Prevent duplicate open suggestions per item
    c.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uniq_reorder_open_by_item
        ON reorder_requests(consumable_id)
        WHERE status IN ('pending','suggested')
    """)

    conn.commit()
    release_connection(conn)

def init_db():
    """Initialize the database schema.
    
    Creates all required tables and performs any necessary migrations.
    Returns True on success, False on failure.
    """
    try:
        logger.info("🚀 Starting database initialization...")
        if not DATABASE_URL:
            logger.error("❌ DATABASE_URL not configured!")
            return False
        # Ensure dirs
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        os.makedirs(INVENTORY_IMAGE_FOLDER, exist_ok=True)

        # Core system tables
        create_basic_tables()

        # Inventory and related
        create_inventory_tables()
        create_inventory_alerts_table()
        create_job_consumables_table()
        create_additional_inventory_tables()
        create_purchasing_integration_tables()
        create_inventory_metrics_tables()  # New enhanced metrics tables
        create_reorder_tables()
        create_default_templates()

        # Enterprise upgrades
        upgrade_user_table_for_enterprise()
        upgrade_equipment_instances_table()
        upgrade_consumables_table()

        insert_sample_data()

        logger.info("🎉 Database initialization completed successfully!")
        return True
    except Exception as e:
        logger.error(f"❌ Database initialization failed: {str(e)}")
        return False

# User class for Flask-Login
class User(UserMixin):
    def __init__(self, identifier, by_username=False):
        conn = get_database_connection()
        c = conn.cursor()
        
        try:
            if by_username:
                c.execute("SELECT id, username, role FROM users WHERE username = %s", (identifier,))
            else:
                c.execute("SELECT id, username, role FROM users WHERE id = %s", (identifier,))
            
            user = c.fetchone()
            if user:
                self.id = str(user[0])
                self.username = user[1]
                self.role = user[2]
            else:
                raise ValueError(f"User with {'username' if by_username else 'id'} {identifier} not found")
        finally:
            release_connection(conn)

    def is_admin(self):
        return self.role == 'admin'

def upgrade_consumables_table_for_separation():
    """Enhanced consumables table with granular separation and purchasing integration"""
    conn = get_connection()
    c = conn.cursor()
    
    # Check existing columns
    c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'consumables'")
    existing_columns = [row[0] for row in c.fetchall()]
    
    # Add new columns for purchasing integration and image storage
    new_columns = [
        ("supplier_sku", "TEXT"),  # Supplier's SKU/part number
        ("internal_sku", "TEXT UNIQUE"),  # Our internal SKU
        ("image_path", "TEXT"),  # Path to uploaded image
        ("image_filename", "TEXT"),  # Original filename of image
        ("barcode", "TEXT"),  # Barcode for scanning
        ("bin_location", "TEXT"),  # Specific bin/shelf location
        ("cost_per_unit", "DECIMAL(10,4)"),  # Cost per unit
        ("last_purchase_date", "TIMESTAMP"),  # When last purchased
        ("last_purchase_cost", "DECIMAL(10,4)"),  # Last purchase cost
        ("preferred_supplier", "BOOLEAN DEFAULT FALSE"),  # If this supplier is preferred for this item
        ("minimum_stock_level", "INTEGER"),  # Minimum stock to maintain
        ("maximum_stock_level", "INTEGER"),  # Maximum stock to maintain
        ("weight_per_unit", "DECIMAL(10,4)"),  # Weight per unit for shipping
        ("dimensions", "JSONB"),  # Dimensions (length, width, height)
    ]
    
    for col_name, col_type in new_columns:
        if col_name not in existing_columns:
            try:
                c.execute(f"ALTER TABLE consumables ADD COLUMN {col_name} {col_type}")
            except psycopg2.Error as e:
                if "already exists" not in str(e):
                    logging.error(f"Error adding column {col_name}: {str(e)}")
    
    # Create auto-generated internal SKU for existing items
    c.execute("""
        UPDATE consumables 
        SET internal_sku = 'SKU-' || LPAD(id::text, 6, '0') 
        WHERE internal_sku IS NULL
    """)
    
    # Update the unique constraint to be more granular
    c.execute("DROP INDEX IF EXISTS idx_consumables_unique_item")
    
    # Create new granular unique constraint - each combination of key specs + supplier should be unique
    c.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_consumables_granular_unique 
        ON consumables(category, base_name, specifications, supplier) 
        WHERE active = TRUE
    """)
    
    # Create indexes for new fields
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_supplier_sku ON consumables(supplier_sku)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_internal_sku ON consumables(internal_sku)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_bin_location ON consumables(bin_location)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_cost ON consumables(cost_per_unit)")
    
    conn.commit()
    release_connection(conn)

def update_item_templates_for_separation():
    """Update item templates to enforce proper separation"""
    conn = get_connection()
    c = conn.cursor()
    
    # Clear existing templates and add enhanced ones
    c.execute("DELETE FROM item_templates")
    
    enhanced_templates = [
        # Enhanced Bolts template with mandatory separation
        {
            'category': 'fasteners',
            'base_name': 'bolt',
            'required_specs': {
                'diameter': {
                    'type': 'string', 
                    'label': 'Diameter',
                    'values': ['1/4"', '5/16"', '3/8"', '7/16"', '1/2"', '9/16"', '5/8"', '3/4"', '7/8"', '1"', '1-1/8"', '1-1/4"'],
                    'separation_key': True  # This field creates separate inventory items
                },
                'length': {
                    'type': 'string', 
                    'label': 'Length',
                    'pattern': r'^\d+(\.\d+)?(\s)?("|in|inch|inches)?$',
                    'placeholder': 'e.g., 2", 3.5", 4 inches',
                    'separation_key': True
                },
                'material': {
                    'type': 'string', 
                    'label': 'Material',
                    'values': ['steel', 'stainless_steel_304', 'stainless_steel_316', 'galvanized', 'zinc_plated', 'brass', 'aluminum'],
                    'separation_key': True
                },
                'thread_type': {
                    'type': 'string', 
                    'label': 'Thread Type',
                    'values': ['coarse', 'fine'],
                    'separation_key': True
                }
            },
            'optional_specs': {
                'head_type': {
                    'type': 'string', 
                    'label': 'Head Type',
                    'values': ['hex', 'carriage', 'socket', 'button', 'flat', 'round']
                },
                'drive_type': {
                    'type': 'string',
                    'label': 'Drive Type', 
                    'values': ['hex', 'phillips', 'flathead', 'torx', 'square']
                },
                'finish': {
                    'type': 'string',
                    'label': 'Finish',
                    'values': ['plain', 'zinc_plated', 'hot_dip_galvanized', 'black_oxide']
                }
            },
            'default_unit': 'units',
            'naming_pattern': '{diameter} x {length} {material} {thread_type} {head_type} bolt',
            'description': 'Bolts separated by diameter, length, material, and thread type. Each combination creates a unique inventory item.',
            'separation_rule': 'Each unique combination of diameter, length, material, thread_type, and supplier creates a separate inventory item'
        },
        
        # Enhanced Cable template with mandatory separation  
        {
            'category': 'electrical',
            'base_name': 'cable',
            'required_specs': {
                'gauge_size': {
                    'type': 'string',
                    'label': 'Wire Gauge/Size', 
                    'values': ['12 AWG', '14 AWG', '16 AWG', '18 AWG', '20 AWG', '22 AWG', '6mm', '8mm', '10mm', '12mm', '16mm', '20mm'],
                    'separation_key': True
                },
                'conductor_material': {
                    'type': 'string', 
                    'label': 'Conductor Material',
                    'values': ['copper', 'aluminum', 'copper_clad_aluminum', 'steel'],
                    'separation_key': True
                },
                'construction': {
                    'type': 'string', 
                    'label': 'Construction',
                    'values': ['solid', 'stranded', 'braided', 'twisted_pair'],
                    'separation_key': True
                },
                'insulation_type': {
                    'type': 'string', 
                    'label': 'Insulation',
                    'values': ['pvc', 'xlpe', 'rubber', 'teflon', 'pe', 'bare'],
                    'separation_key': True
                }
            },
            'optional_specs': {
                'voltage_rating': {
                    'type': 'string', 
                    'label': 'Voltage Rating',
                    'values': ['300V', '600V', '1000V', '5kV', '15kV', '25kV', '35kV']
                },
                'conductor_count': {
                    'type': 'integer', 
                    'label': 'Number of Conductors',
                    'min': 1, 
                    'max': 50
                },
                'shielding': {
                    'type': 'string',
                    'label': 'Shielding',
                    'values': ['none', 'foil', 'braid', 'spiral']
                },
                'jacket_color': {
                    'type': 'string',
                    'label': 'Jacket Color',
                    'values': ['black', 'white', 'red', 'blue', 'green', 'yellow', 'orange', 'purple']
                }
            },
            'default_unit': 'feet',
            'naming_pattern': '{gauge_size} {conductor_material} {construction} {insulation_type} cable',
            'description': 'Electrical cables separated by gauge, conductor material, construction, and insulation. Each combination creates a unique inventory item.',
            'separation_rule': 'Each unique combination of gauge_size, conductor_material, construction, insulation_type, and supplier creates a separate inventory item'
        },
        
        # Enhanced Safety Equipment template
        {
            'category': 'safety',
            'base_name': 'gloves',
            'required_specs': {
                'material': {
                    'type': 'string', 
                    'label': 'Material',
                    'values': ['rubber', 'leather', 'nitrile', 'latex', 'vinyl', 'cotton'],
                    'separation_key': True
                },
                'protection_class': {
                    'type': 'string', 
                    'label': 'Protection Class',
                    'values': ['class_0', 'class_1', 'class_2', 'class_3', 'class_4', 'non_electrical'],
                    'separation_key': True
                },
                'size': {
                    'type': 'string', 
                    'label': 'Size',
                    'values': ['XS', 'S', 'M', 'L', 'XL', 'XXL'],
                    'separation_key': True
                }
            },
            'optional_specs': {
                'cuff_length': {
                    'type': 'string', 
                    'label': 'Cuff Length',
                    'values': ['short', 'medium', 'long']
                },
                'color': {
                    'type': 'string',
                    'label': 'Color',
                    'values': ['black', 'orange', 'red', 'yellow', 'white', 'blue']
                },
                'thickness': {
                    'type': 'string',
                    'label': 'Thickness',
                    'values': ['light', 'medium', 'heavy']
                }
            },
            'default_unit': 'pairs',
            'naming_pattern': '{material} {protection_class} {size} safety gloves',
            'description': 'Safety gloves separated by material, protection class, and size. Each combination creates a unique inventory item.',
            'separation_rule': 'Each unique combination of material, protection_class, size, and supplier creates a separate inventory item'
        },
        
        # Add more categories as needed
        {
            'category': 'hardware',
            'base_name': 'washer',
            'required_specs': {
                'diameter': {
                    'type': 'string',
                    'label': 'Diameter',
                    'values': ['1/4"', '5/16"', '3/8"', '7/16"', '1/2"', '9/16"', '5/8"', '3/4"', '1"'],
                    'separation_key': True
                },
                'material': {
                    'type': 'string',
                    'label': 'Material', 
                    'values': ['steel', 'stainless_steel', 'galvanized', 'brass', 'nylon'],
                    'separation_key': True
                },
                'type': {
                    'type': 'string',
                    'label': 'Washer Type',
                    'values': ['flat', 'lock', 'spring', 'fender'],
                    'separation_key': True
                }
            },
            'optional_specs': {
                'thickness': {
                    'type': 'string',
                    'label': 'Thickness',
                    'pattern': r'^\d+(\.\d+)?(mm|"|in|inch)?$'
                }
            },
            'default_unit': 'units',
            'naming_pattern': '{diameter} {material} {type} washer',
            'description': 'Washers separated by diameter, material, and type.',
            'separation_rule': 'Each unique combination of diameter, material, type, and supplier creates a separate inventory item'
        }
    ]
    
    for template in enhanced_templates:
        c.execute("""
            INSERT INTO item_templates 
            (category, base_name, required_specs, optional_specs, default_unit, naming_pattern, description)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            template['category'],
            template['base_name'],
            json.dumps(template['required_specs']),
            json.dumps(template.get('optional_specs', {})),
            template['default_unit'],
            template['naming_pattern'],
            template['description']
        ))
    
    conn.commit()
    release_connection(conn)

def create_inventory_metrics_tables():
    """Create tables for enhanced inventory metrics and analytics"""
    conn = get_connection()
    c = conn.cursor()

    # Supplier price history for tracking price changes over time
    c.execute('''CREATE TABLE IF NOT EXISTS supplier_price_history (
        id SERIAL PRIMARY KEY,
        supplier_item_id INTEGER NOT NULL,
        consumable_id INTEGER NOT NULL,
        supplier_id INTEGER,
        old_price DECIMAL(10,4),
        new_price DECIMAL(10,4) NOT NULL,
        change_percentage DECIMAL(8,4),
        change_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        change_source TEXT DEFAULT 'manual',  -- manual, api_sync, invoice, quote
        recorded_by INTEGER,
        notes TEXT,
        FOREIGN KEY (supplier_item_id) REFERENCES supplier_items(id) ON DELETE CASCADE,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL,
        FOREIGN KEY (recorded_by) REFERENCES users(id) ON DELETE SET NULL
    )''')

    # Purchase order items for line-item tracking
    c.execute('''CREATE TABLE IF NOT EXISTS purchase_order_items (
        id SERIAL PRIMARY KEY,
        purchase_order_id INTEGER NOT NULL,
        consumable_id INTEGER NOT NULL,
        quantity_ordered INTEGER NOT NULL,
        quantity_received INTEGER DEFAULT 0,
        unit_price DECIMAL(10,4),
        line_total DECIMAL(12,2),
        received_date TIMESTAMP,
        condition_notes TEXT,
        FOREIGN KEY (purchase_order_id) REFERENCES purchase_orders(id) ON DELETE CASCADE,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE
    )''')

    # Supplier contacts for multiple contacts per supplier
    c.execute('''CREATE TABLE IF NOT EXISTS supplier_contacts (
        id SERIAL PRIMARY KEY,
        supplier_id INTEGER NOT NULL,
        contact_type TEXT DEFAULT 'primary',  -- primary, billing, sales, support, emergency
        name TEXT NOT NULL,
        title TEXT,
        email TEXT,
        phone TEXT,
        mobile TEXT,
        preferred_contact_method TEXT DEFAULT 'email',  -- email, phone, portal
        notes TEXT,
        active BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
    )''')

    # Shipment tracking for incoming orders
    c.execute('''CREATE TABLE IF NOT EXISTS shipment_tracking (
        id SERIAL PRIMARY KEY,
        purchase_order_id INTEGER NOT NULL,
        carrier TEXT,
        tracking_number TEXT,
        status TEXT DEFAULT 'pending',  -- pending, in_transit, out_for_delivery, delivered, exception
        estimated_arrival TIMESTAMP,
        actual_arrival TIMESTAMP,
        ship_date TIMESTAMP,
        last_update TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        current_location TEXT,
        delivery_confirmed BOOLEAN DEFAULT FALSE,
        delivery_confirmed_by INTEGER,
        delivery_notes TEXT,
        signature_captured BOOLEAN DEFAULT FALSE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (purchase_order_id) REFERENCES purchase_orders(id) ON DELETE CASCADE,
        FOREIGN KEY (delivery_confirmed_by) REFERENCES users(id) ON DELETE SET NULL
    )''')

    # Inventory locations for multi-location support
    c.execute('''CREATE TABLE IF NOT EXISTS inventory_locations (
        id SERIAL PRIMARY KEY,
        name TEXT NOT NULL UNIQUE,
        location_type TEXT DEFAULT 'warehouse',  -- warehouse, job_site, truck, office
        address TEXT,
        city TEXT,
        state TEXT,
        zip_code TEXT,
        manager_id INTEGER,
        is_primary BOOLEAN DEFAULT FALSE,
        active BOOLEAN DEFAULT TRUE,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (manager_id) REFERENCES users(id) ON DELETE SET NULL
    )''')

    # Inventory transfers between locations
    c.execute('''CREATE TABLE IF NOT EXISTS inventory_transfers (
        id SERIAL PRIMARY KEY,
        from_location_id INTEGER NOT NULL,
        to_location_id INTEGER NOT NULL,
        consumable_id INTEGER NOT NULL,
        quantity INTEGER NOT NULL,
        status TEXT DEFAULT 'requested',  -- requested, approved, in_transit, received, cancelled
        requested_by INTEGER,
        approved_by INTEGER,
        request_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        approval_date TIMESTAMP,
        ship_date TIMESTAMP,
        received_date TIMESTAMP,
        tracking_info TEXT,
        notes TEXT,
        FOREIGN KEY (from_location_id) REFERENCES inventory_locations(id) ON DELETE CASCADE,
        FOREIGN KEY (to_location_id) REFERENCES inventory_locations(id) ON DELETE CASCADE,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
        FOREIGN KEY (requested_by) REFERENCES users(id) ON DELETE SET NULL,
        FOREIGN KEY (approved_by) REFERENCES users(id) ON DELETE SET NULL
    )''')

    # Inventory forecasts for demand predictions
    c.execute('''CREATE TABLE IF NOT EXISTS inventory_forecasts (
        id SERIAL PRIMARY KEY,
        consumable_id INTEGER NOT NULL,
        forecast_date DATE NOT NULL,
        forecast_period TEXT DEFAULT 'monthly',  -- daily, weekly, monthly, quarterly
        predicted_quantity DECIMAL(10,2),
        confidence_level DECIMAL(5,4),  -- 0.0 to 1.0
        based_on_data_points INTEGER,
        model_used TEXT DEFAULT 'moving_average',  -- moving_average, exponential_smoothing, ml_model
        actual_quantity DECIMAL(10,2),  -- filled in after the period passes
        variance DECIMAL(10,2),  -- actual - predicted
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
        UNIQUE(consumable_id, forecast_date, forecast_period)
    )''')

    # Supplier API sync log
    c.execute('''CREATE TABLE IF NOT EXISTS supplier_api_sync_log (
        id SERIAL PRIMARY KEY,
        supplier_id INTEGER NOT NULL,
        sync_type TEXT NOT NULL,  -- catalog, prices, availability, orders
        status TEXT DEFAULT 'pending',  -- pending, running, completed, failed
        started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        completed_at TIMESTAMP,
        items_synced INTEGER DEFAULT 0,
        items_failed INTEGER DEFAULT 0,
        error_message TEXT,
        sync_details JSONB,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
    )''')

    # Inventory metrics snapshots for historical tracking
    c.execute('''CREATE TABLE IF NOT EXISTS inventory_metrics_snapshots (
        id SERIAL PRIMARY KEY,
        snapshot_date DATE NOT NULL,
        total_items INTEGER,
        total_value DECIMAL(14,2),
        items_in_stock INTEGER,
        items_low_stock INTEGER,
        items_out_of_stock INTEGER,
        total_categories INTEGER,
        avg_turnover_ratio DECIMAL(8,4),
        fill_rate DECIMAL(5,4),
        stockout_rate DECIMAL(5,4),
        metrics_data JSONB,  -- additional detailed metrics
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(snapshot_date)
    )''')

    # Create indexes for performance
    c.execute("CREATE INDEX IF NOT EXISTS idx_price_history_consumable ON supplier_price_history(consumable_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_price_history_date ON supplier_price_history(change_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_po_items_order ON purchase_order_items(purchase_order_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_shipment_status ON shipment_tracking(status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_shipment_po ON shipment_tracking(purchase_order_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_transfers_status ON inventory_transfers(status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_forecasts_date ON inventory_forecasts(forecast_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_api_sync_supplier ON supplier_api_sync_log(supplier_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_metrics_date ON inventory_metrics_snapshots(snapshot_date)")

    conn.commit()
    release_connection(conn)
    logger.info("✅ Inventory metrics tables created successfully")

def create_purchasing_integration_tables():
    """Create tables for purchasing software integration"""
    conn = get_connection()
    c = conn.cursor()

    # Suppliers table
    c.execute('''CREATE TABLE IF NOT EXISTS suppliers (
        id SERIAL PRIMARY KEY,
        name TEXT NOT NULL UNIQUE,
        contact_person TEXT,
        email TEXT,
        phone TEXT,
        address TEXT,
        city TEXT,
        state TEXT,
        zip_code TEXT,
        country TEXT DEFAULT 'USA',
        website TEXT,
        account_number TEXT,
        payment_terms TEXT,
        shipping_terms TEXT,
        preferred BOOLEAN DEFAULT FALSE,
        active BOOLEAN DEFAULT TRUE,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Add API fields here
    c.execute("ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS api_endpoint TEXT");
    c.execute("ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS api_key TEXT");
    c.execute("ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS api_username TEXT");
    c.execute("ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS api_password TEXT");
    c.execute("ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS api_auth_type TEXT");
    
    # Enhanced supplier_items table with more purchasing details
    c.execute("DROP TABLE IF EXISTS supplier_items CASCADE")
    c.execute('''CREATE TABLE IF NOT EXISTS supplier_items (
        id SERIAL PRIMARY KEY,
        consumable_id INTEGER NOT NULL,
        supplier_id INTEGER,
        supplier_name TEXT NOT NULL,
        supplier_sku TEXT,
        supplier_part_number TEXT,
        supplier_description TEXT,
        cost_per_unit DECIMAL(10,4),
        minimum_order_quantity INTEGER DEFAULT 1,
        order_multiple INTEGER DEFAULT 1,
        lead_time_days INTEGER,
        preferred BOOLEAN DEFAULT FALSE,
        current_supplier BOOLEAN DEFAULT FALSE,
        last_ordered TIMESTAMP,
        last_cost DECIMAL(10,4),
        notes TEXT,
        catalog_page TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
        UNIQUE(consumable_id, supplier_name, supplier_sku)
    )''')
    
    # Purchase orders table for future integration
    c.execute('''CREATE TABLE IF NOT EXISTS purchase_orders (
        id SERIAL PRIMARY KEY,
        po_number TEXT UNIQUE NOT NULL,
        supplier_id INTEGER NOT NULL,
        status TEXT DEFAULT 'draft',
        order_date DATE,
        expected_delivery_date DATE,
        actual_delivery_date DATE,
        subtotal DECIMAL(10,2),
        tax_amount DECIMAL(10,2),
        shipping_cost DECIMAL(10,2),
        total_amount DECIMAL(10,2),
        notes TEXT,
        created_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
        FOREIGN KEY (created_by) REFERENCES users(id)
    )''')
    
    # Purchase order items
    c.execute('''CREATE TABLE IF NOT EXISTS purchase_order_items (
        id SERIAL PRIMARY KEY,
        po_id INTEGER NOT NULL,
        consumable_id INTEGER NOT NULL,
        supplier_item_id INTEGER,
        quantity INTEGER NOT NULL,
        unit_cost DECIMAL(10,4),
        total_cost DECIMAL(10,2),
        received_quantity INTEGER DEFAULT 0,
        notes TEXT,
        FOREIGN KEY (po_id) REFERENCES purchase_orders(id) ON DELETE CASCADE,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id),
        FOREIGN KEY (supplier_item_id) REFERENCES supplier_items(id)
    )''')
    
    # Inventory transactions for tracking all movements
    c.execute('''CREATE TABLE IF NOT EXISTS inventory_transactions (
        id SERIAL PRIMARY KEY,
        consumable_id INTEGER NOT NULL,
        transaction_type TEXT NOT NULL, -- 'purchase', 'usage', 'adjustment', 'transfer'
        quantity_change INTEGER NOT NULL, -- positive for additions, negative for usage
        quantity_before INTEGER,
        quantity_after INTEGER,
        reference_id INTEGER, -- Could be PO ID, Job ID, etc.
        reference_type TEXT, -- 'purchase_order', 'job', 'adjustment'
        cost_per_unit DECIMAL(10,4),
        total_cost DECIMAL(10,2),
        notes TEXT,
        performed_by INTEGER,
        transaction_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id),
        FOREIGN KEY (performed_by) REFERENCES users(id)
    )''')
    
    # Create indexes
    c.execute("CREATE INDEX IF NOT EXISTS idx_supplier_items_consumable ON supplier_items(consumable_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_supplier_items_supplier ON supplier_items(supplier_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_supplier_items_sku ON supplier_items(supplier_sku)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_purchase_orders_supplier ON purchase_orders(supplier_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_purchase_orders_status ON purchase_orders(status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_transactions_consumable ON inventory_transactions(consumable_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_transactions_type ON inventory_transactions(transaction_type)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_transactions_date ON inventory_transactions(transaction_date)")
    
    conn.commit()
    release_connection(conn)

# Enhanced Inventory Name Generator with stricter separation
class EnhancedInventoryNameGenerator:
    @staticmethod
    def generate_standardized_name(category, base_name, specifications):
        """Generate a standardized name with all separation keys included"""
        if category == 'fasteners' and base_name == 'bolt':
            # Must include ALL separation keys for bolts
            parts = []
            parts.append(specifications.get('diameter', ''))
            parts.append(f"x {specifications.get('length', '')}")
            parts.append(specifications.get('material', ''))
            parts.append(specifications.get('thread_type', ''))
            if specifications.get('head_type'):
                parts.append(specifications.get('head_type'))
            parts.append('bolt')
            return ' '.join([p for p in parts if p and p != 'x ']).strip()
        
        elif category == 'electrical' and base_name == 'cable':
            # Must include ALL separation keys for cables
            parts = []
            parts.append(specifications.get('gauge_size', ''))
            parts.append(specifications.get('conductor_material', ''))
            parts.append(specifications.get('construction', ''))
            parts.append(specifications.get('insulation_type', ''))
            parts.append('cable')
            return ' '.join([p for p in parts if p]).strip()
        
        elif category == 'safety' and base_name == 'gloves':
            # Must include ALL separation keys for gloves
            parts = []
            parts.append(specifications.get('material', ''))
            parts.append(specifications.get('protection_class', ''))
            parts.append(specifications.get('size', ''))
            parts.extend(['safety', 'gloves'])
            return ' '.join([p for p in parts if p]).strip()
        
        elif category == 'hardware' and base_name == 'washer':
            parts = []
            parts.append(specifications.get('diameter', ''))
            parts.append(specifications.get('material', ''))
            parts.append(specifications.get('type', ''))
            parts.append('washer')
            return ' '.join([p for p in parts if p]).strip()
        
        # Fallback - use all specifications
        spec_parts = []
        for key, value in specifications.items():
            if value:
                spec_parts.append(str(value))
        spec_parts.append(base_name)
        return ' '.join(spec_parts).strip()
    
    @staticmethod
    def generate_separation_key(category, base_name, specifications, supplier):
        """Generate a key that enforces separation by critical attributes"""
        separation_parts = [category, base_name]
        
        # Get template to know which specs are separation keys
        template = get_template_for_item(category, base_name)
        if template:
            required_specs = template.get('required_specs', {})
            for spec_name, spec_config in required_specs.items():
                if spec_config.get('separation_key', False):
                    value = specifications.get(spec_name, '')
                    if value:
                        normalized_value = str(value).lower().replace(' ', '_').replace('/', '_').replace('"', 'in')
                        separation_parts.append(f"{spec_name}:{normalized_value}")
        
        # Always include supplier in separation
        separation_parts.append(f"supplier:{supplier.lower().replace(' ', '_')}")
        
        return '|'.join(separation_parts)
    
    @staticmethod
    def generate_internal_sku(category, base_name, specifications, sequence_number):
        """Generate internal SKU with category prefix"""
        category_prefixes = {
            'fasteners': 'FAST',
            'electrical': 'ELEC', 
            'safety': 'SAFE',
            'hardware': 'HARD',
            'tools': 'TOOL'
        }
        
        prefix = category_prefixes.get(category, 'MISC')
        return f"{prefix}-{sequence_number:06d}"

def get_template_for_item(category, base_name):
    """Get template for specific category and base name"""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT required_specs, optional_specs FROM item_templates WHERE category = %s AND base_name = %s", 
              (category, base_name))
    result = c.fetchone()
    release_connection(conn)
    
    if result:
        return {
            'required_specs': json.loads(result[0]),
            'optional_specs': json.loads(result[1]) if result[1] else {}
        }
    return None

# Add these functions to init_db.py main function
def init_db():
    try:
        conn = get_connection()
        c = conn.cursor()

        # Existing table creation code...
        # [Keep all your existing init_db code]
        
        # Add new enhancements
        upgrade_consumables_table_for_separation()
        update_item_templates_for_separation()
        create_purchasing_integration_tables()
        
        # Rest of existing init_db code...
        
        conn.commit()
        logging.info("Enhanced database initialization with granular separation completed successfully")
    except Exception as e:
        logging.error(f"Database initialization failed: {str(e)}")
        raise
    finally:
        release_connection(conn)        

def create_inventory_tables():
    """Create new inventory-related tables"""
    conn = get_database_connection()
    c = conn.cursor()
    
    try:
        logger.info("🔧 Creating inventory management tables...")
        
        # Inventory alerts table
        c.execute('''CREATE TABLE IF NOT EXISTS inventory_alerts (
            id SERIAL PRIMARY KEY,
            consumable_id INTEGER NOT NULL,
            alert_type TEXT NOT NULL,
            alert_level TEXT NOT NULL,
            message TEXT NOT NULL,
            acknowledged BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            acknowledged_at TIMESTAMP,
            acknowledged_by INTEGER,
            FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE
        )''')
        logger.info("  ✅ Created inventory_alerts table")
        
        # Job consumables tracking
        c.execute('''CREATE TABLE IF NOT EXISTS job_consumables (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL,
            consumable_id INTEGER NOT NULL,
            quantity_used DECIMAL(10,4) NOT NULL,
            cost_per_unit DECIMAL(10,4),
            total_cost DECIMAL(10,2),
            date_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            recorded_by INTEGER,
            FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE,
            FOREIGN KEY (consumable_id) REFERENCES consumables(id)
        )''')
        logger.info("  ✅ Created job_consumables table")
        
        # Inventory transactions
        c.execute('''CREATE TABLE IF NOT EXISTS inventory_transactions (
            id SERIAL PRIMARY KEY,
            consumable_id INTEGER NOT NULL,
            transaction_type TEXT NOT NULL,
            quantity_change INTEGER NOT NULL,
            quantity_before INTEGER,
            quantity_after INTEGER,
            cost_per_unit DECIMAL(10,4),
            total_cost DECIMAL(10,2),
            notes TEXT,
            performed_by INTEGER,
            transaction_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (consumable_id) REFERENCES consumables(id)
        )''')
        logger.info("  ✅ Created inventory_transactions table")
        
        # Suppliers table
        c.execute('''CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            contact_person TEXT,
            email TEXT,
            phone TEXT,
            address TEXT,
            active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        logger.info("  ✅ Created suppliers table")
        
        # Supplier items
        c.execute('''CREATE TABLE IF NOT EXISTS supplier_items (
            id SERIAL PRIMARY KEY,
            consumable_id INTEGER NOT NULL,
            supplier_id INTEGER,
            supplier_name TEXT NOT NULL,
            supplier_sku TEXT,
            supplier_part_number TEXT,
            cost_per_unit DECIMAL(10,4),
            lead_time_days INTEGER,
            minimum_order_quantity INTEGER DEFAULT 1,
            preferred BOOLEAN DEFAULT FALSE,
            current_supplier BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
        )''')
        logger.info("  ✅ Created supplier_items table")
        
        # Item templates
        c.execute('''CREATE TABLE IF NOT EXISTS item_templates (
            id SERIAL PRIMARY KEY,
            category TEXT NOT NULL,
            base_name TEXT NOT NULL,
            required_specs JSONB NOT NULL,
            optional_specs JSONB,
            default_unit TEXT,
            naming_pattern TEXT NOT NULL,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(category, base_name)
        )''')
        logger.info("  ✅ Created item_templates table")
        
        # Create indexes
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_inventory_alerts_consumable ON inventory_alerts(consumable_id)",
            "CREATE INDEX IF NOT EXISTS idx_inventory_alerts_acknowledged ON inventory_alerts(acknowledged)",
            "CREATE INDEX IF NOT EXISTS idx_job_consumables_job ON job_consumables(job_id)",
            "CREATE INDEX IF NOT EXISTS idx_job_consumables_consumable ON job_consumables(consumable_id)",
            "CREATE INDEX IF NOT EXISTS idx_inventory_transactions_consumable ON inventory_transactions(consumable_id)",
            "CREATE INDEX IF NOT EXISTS idx_inventory_transactions_type ON inventory_transactions(transaction_type)",
            "CREATE INDEX IF NOT EXISTS idx_supplier_items_consumable ON supplier_items(consumable_id)",
            "CREATE INDEX IF NOT EXISTS idx_supplier_items_supplier ON supplier_items(supplier_id)"
        ]
        
        for index_sql in indexes:
            c.execute(index_sql)
        
        logger.info("✅ All inventory table indexes created")
        
        conn.commit()
        logger.info("✅ All inventory tables created successfully")
        
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error creating inventory tables: {e}")
        raise
    finally:
        release_connection(conn)

def create_default_templates():
    """Create default item templates for construction items"""
    conn = get_database_connection()
    c = conn.cursor()
    
    try:
        logger.info("🔧 Creating default item templates...")
        # Ensure the item_templates table exists before inserting
        c.execute('''CREATE TABLE IF NOT EXISTS item_templates (
            id SERIAL PRIMARY KEY,
            category TEXT NOT NULL,
            base_name TEXT NOT NULL,
            required_specs JSONB NOT NULL,
            optional_specs JSONB,
            default_unit TEXT,
            naming_pattern TEXT NOT NULL,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(category, base_name)
        )''')
        
        templates = [
            {
                'category': 'fasteners',
                'base_name': 'bolt',
                'required_specs': {
                    'diameter': {'type': 'select', 'values': ['1/4"', '3/8"', '1/2"', '5/8"', '3/4"', '1"'], 'label': 'Diameter'},
                    'length': {'type': 'text', 'label': 'Length'},
                    'material': {'type': 'select', 'values': ['steel', 'stainless_steel', 'galvanized'], 'label': 'Material'},
                    'grade': {'type': 'select', 'values': ['2', '5', '8'], 'label': 'Grade'}
                },
                'optional_specs': {
                    'head_type': {'type': 'select', 'values': ['hex', 'carriage', 'socket'], 'label': 'Head Type'}
                },
                'default_unit': 'each',
                'naming_pattern': '{diameter} x {length} {material} Grade {grade} Bolt',
                'description': 'Standard construction bolts with diameter, length, material, and grade specifications'
            },
            {
                'category': 'electrical',
                'base_name': 'cable',
                'required_specs': {
                    'gauge': {'type': 'select', 'values': ['12', '14', '16', '18'], 'label': 'Wire Gauge'},
                    'conductor_material': {'type': 'select', 'values': ['copper', 'aluminum'], 'label': 'Conductor Material'},
                    'insulation': {'type': 'select', 'values': ['THHN', 'THWN'], 'label': 'Insulation Type'}
                },
                'optional_specs': {
                    'voltage_rating': {'type': 'select', 'values': ['300V', '600V', '1000V'], 'label': 'Voltage Rating'}
                },
                'default_unit': 'feet',
                'naming_pattern': '{gauge} AWG {conductor_material} {insulation} Cable',
                'description': 'Electrical cables with gauge, conductor material, and insulation specifications'
            },
            {
                'category': 'safety',
                'base_name': 'gloves',
                'required_specs': {
                    'material': {'type': 'select', 'values': ['leather', 'rubber', 'nitrile'], 'label': 'Material'},
                    'size': {'type': 'select', 'values': ['S', 'M', 'L', 'XL'], 'label': 'Size'},
                    'protection_class': {'type': 'select', 'values': ['general', 'cut_resistant', 'electrical'], 'label': 'Protection Class'}
                },
                'optional_specs': {
                    'color': {'type': 'text', 'label': 'Color'}
                },
                'default_unit': 'pairs',
                'naming_pattern': '{material} {protection_class} {size} Safety Gloves',
                'description': 'Safety gloves with material, size, and protection class specifications'
            }
        ]
        
        for template in templates:
            c.execute("""
                INSERT INTO item_templates 
                (category, base_name, required_specs, optional_specs, default_unit, naming_pattern, description)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (category, base_name) DO UPDATE SET
                    required_specs = EXCLUDED.required_specs,
                    optional_specs = EXCLUDED.optional_specs,
                    default_unit = EXCLUDED.default_unit,
                    naming_pattern = EXCLUDED.naming_pattern,
                    description = EXCLUDED.description
            """, (
                template['category'],
                template['base_name'],
                json.dumps(template['required_specs']),
                json.dumps(template.get('optional_specs', {})),
                template['default_unit'],
                template['naming_pattern'],
                template['description']
            ))
        
        conn.commit()
        logger.info(f"✅ Created {len(templates)} default templates")
        
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error creating templates: {e}")
        raise
    finally:
        release_connection(conn)

def create_basic_tables():
    conn = None
    try:
        conn = get_connection()
        conn.autocommit = True
        c = conn.cursor()

        # Users table with additional fields
        c.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT NOT NULL UNIQUE,
                full_name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                job_title TEXT,
                role TEXT DEFAULT 'user',
                password TEXT NOT NULL,
                certifications JSONB DEFAULT '[]'::JSONB,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Add missing columns idempotently
        c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS address TEXT")
        c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS home_phone TEXT")
        c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS cell_phone TEXT")
        c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS work_phone TEXT")

        # Certifications table for options
        c.execute('''
            CREATE TABLE IF NOT EXISTS certifications (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                issuing_company TEXT,
                validity_term_years INTEGER,
                requirements_text TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Insert preset certifications if needed
        preset_certs = [
            ('forklift', 'Forklift Operator'),
            ('commercial_driver', 'Commercial Driver License'),
            ('boom_operator', 'Boom Truck Operator'),
            ('crane_certified', 'Crane Operator Certified')
        ]
        for name, desc in preset_certs:
            c.execute('''
                INSERT INTO certifications (name, description)
                VALUES (%s, %s)
                ON CONFLICT (name) DO NOTHING
            ''', (name, desc))

    except Exception as e:
        logging.error(f"Error creating basic tables: {str(e)}")
    finally:
        try:
            if conn:
                release_connection(conn)
        except Exception:
            pass

def create_inventory_alerts_table():
    """Create inventory alerts table for notifications"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS inventory_alerts (
        id SERIAL PRIMARY KEY,
        consumable_id INTEGER NOT NULL,
        alert_type TEXT NOT NULL, -- 'low_stock', 'out_of_stock', 'reorder_needed'
        alert_level TEXT NOT NULL, -- 'info', 'warning', 'critical'
        message TEXT NOT NULL,
        acknowledged BOOLEAN DEFAULT FALSE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        acknowledged_at TIMESTAMP,
        acknowledged_by INTEGER,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
        FOREIGN KEY (acknowledged_by) REFERENCES users(id)
    )''')
    
    # Create indexes
    c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_alerts_consumable ON inventory_alerts(consumable_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_alerts_acknowledged ON inventory_alerts(acknowledged)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_alerts_level ON inventory_alerts(alert_level)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_alerts_created ON inventory_alerts(created_at)")
    
    conn.commit()
    release_connection(conn)

def create_job_consumables_table():
    """Create job consumables tracking table"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS job_consumables (
        id SERIAL PRIMARY KEY,
        job_id INTEGER NOT NULL,
        consumable_id INTEGER NOT NULL,
        quantity_used DECIMAL(10,4) NOT NULL,
        cost_per_unit DECIMAL(10,4),
        total_cost DECIMAL(10,2),
        date_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        notes TEXT,
        recorded_by INTEGER,
        FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE,
        FOREIGN KEY (consumable_id) REFERENCES consumables(id),
        FOREIGN KEY (recorded_by) REFERENCES users(id)
    )''')
    
    # Create indexes for performance
    c.execute("CREATE INDEX IF NOT EXISTS idx_job_consumables_job ON job_consumables(job_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_job_consumables_consumable ON job_consumables(consumable_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_job_consumables_date ON job_consumables(date_used)")
    
    conn.commit()
    release_connection(conn)

def create_additional_inventory_tables():
    """Create all additional inventory-related tables"""
    conn = get_connection()
    c = conn.cursor()
    
    # Update consumables table with missing columns
    c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'consumables'")
    existing_columns = [row[0] for row in c.fetchall()]
    
    # Add missing columns to consumables table
    additional_columns = [
        ("supplier_sku", "TEXT"),
        ("internal_sku", "TEXT"),
        ("image_path", "TEXT"),
        ("image_filename", "TEXT"),
        ("barcode", "TEXT"),
        ("bin_location", "TEXT"),
        ("cost_per_unit", "DECIMAL(10,4)"),
        ("last_purchase_date", "TIMESTAMP"),
        ("last_purchase_cost", "DECIMAL(10,4)"),
        ("preferred_supplier", "BOOLEAN DEFAULT FALSE"),
        ("minimum_stock_level", "INTEGER"),
        ("maximum_stock_level", "INTEGER"),
        ("weight_per_unit", "DECIMAL(10,4)"),
        ("dimensions", "JSONB"),
        ("hazmat", "BOOLEAN DEFAULT FALSE"),
        ("requires_certification", "BOOLEAN DEFAULT FALSE"),
        ("product_url", "TEXT"),
        ("notes", "TEXT")
    ]
    
    for col_name, col_type in additional_columns:
        if col_name not in existing_columns:
            try:
                c.execute(f"ALTER TABLE consumables ADD COLUMN {col_name} {col_type}")
            except psycopg2.Error as e:
                if "already exists" not in str(e):
                    logging.error(f"Error adding column {col_name}: {str(e)}")
    
    # Ensure internal_sku is unique where not null
    c.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_consumables_internal_sku_unique 
        ON consumables(internal_sku) 
        WHERE internal_sku IS NOT NULL
    """)
    
    # Create indexes for performance
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_supplier_sku ON consumables(supplier_sku)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_bin_location ON consumables(bin_location)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_cost ON consumables(cost_per_unit)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_consumables_hazmat ON consumables(hazmat)")
    
    conn.commit()
    release_connection(conn)

def create_pm_tables():
    """Create core Project Management tables (boards, items, columns, docs, widgets, templates, risks)."""
    conn = get_database_connection()
    c = conn.cursor()
    # Boards
    c.execute('''CREATE TABLE IF NOT EXISTS pm_boards (
        id SERIAL PRIMARY KEY,
        name TEXT NOT NULL,
        is_private BOOLEAN DEFAULT FALSE,
        is_shareable BOOLEAN DEFAULT FALSE,
        owner_id INTEGER,
        template_key TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (owner_id) REFERENCES users(id) ON DELETE SET NULL
    )''')
    # Board memberships (permissions)
    c.execute('''CREATE TABLE IF NOT EXISTS pm_board_members (
        id SERIAL PRIMARY KEY,
        board_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        role TEXT DEFAULT 'viewer', -- viewer, editor, admin
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(board_id, user_id),
        FOREIGN KEY (board_id) REFERENCES pm_boards(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    # Columns
    c.execute('''CREATE TABLE IF NOT EXISTS pm_columns (
        id SERIAL PRIMARY KEY,
        board_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        type TEXT NOT NULL, -- text, number, date, timeline, people, status, dropdown, formula, time_tracking, dependency, location, file
        config JSONB,
        position INTEGER DEFAULT 0,
        is_restricted BOOLEAN DEFAULT FALSE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (board_id) REFERENCES pm_boards(id) ON DELETE CASCADE
    )''')
    # Items
    c.execute('''CREATE TABLE IF NOT EXISTS pm_items (
        id SERIAL PRIMARY KEY,
        board_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        status TEXT,
        start_date DATE,
        end_date DATE,
        due_date DATE,
        assignee_id INTEGER,
        position INTEGER DEFAULT 0,
        estimated_hours NUMERIC(10,2),
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (board_id) REFERENCES pm_boards(id) ON DELETE CASCADE,
        FOREIGN KEY (assignee_id) REFERENCES users(id) ON DELETE SET NULL
    )''')
    # Ensure position column exists (idempotent)
    try:
        c.execute("ALTER TABLE pm_items ADD COLUMN IF NOT EXISTS position INTEGER DEFAULT 0")
    except Exception:
        pass
    # Ensure scheduling columns exist (idempotent)
    for col_def in [
        ("start_date", "DATE"),
        ("end_date", "DATE"),
        ("estimated_hours", "NUMERIC(10,2)")
    ]:
        try:
            c.execute(f"ALTER TABLE pm_items ADD COLUMN IF NOT EXISTS {col_def[0]} {col_def[1]}")
        except Exception:
            pass
    # Item values by column
    c.execute('''CREATE TABLE IF NOT EXISTS pm_item_values (
        id SERIAL PRIMARY KEY,
        item_id INTEGER NOT NULL,
        column_id INTEGER NOT NULL,
        value JSONB,
        FOREIGN KEY (item_id) REFERENCES pm_items(id) ON DELETE CASCADE,
        FOREIGN KEY (column_id) REFERENCES pm_columns(id) ON DELETE CASCADE,
        UNIQUE(item_id, column_id)
    )''')
    # Updates/comments per item
    c.execute('''CREATE TABLE IF NOT EXISTS pm_item_updates (
        id SERIAL PRIMARY KEY,
        item_id INTEGER NOT NULL,
        author_id INTEGER,
        content TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (item_id) REFERENCES pm_items(id) ON DELETE CASCADE,
        FOREIGN KEY (author_id) REFERENCES users(id) ON DELETE SET NULL
    )''')
    # Files (references to existing documents table or uploads)
    c.execute('''CREATE TABLE IF NOT EXISTS pm_item_files (
        id SERIAL PRIMARY KEY,
        item_id INTEGER NOT NULL,
        file_path TEXT,
        file_name TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (item_id) REFERENCES pm_items(id) ON DELETE CASCADE
    )''')
    # Workdocs
    c.execute('''CREATE TABLE IF NOT EXISTS pm_docs (
        id SERIAL PRIMARY KEY,
        board_id INTEGER,
        title TEXT NOT NULL,
        content TEXT,
        created_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (board_id) REFERENCES pm_boards(id) ON DELETE CASCADE,
        FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL
    )''')
    # Dashboard widgets (reuse existing table name if present)
    c.execute('''CREATE TABLE IF NOT EXISTS pm_dashboard_widgets (
        id SERIAL PRIMARY KEY,
        user_id INTEGER NOT NULL,
        widget_type TEXT NOT NULL,
        config JSONB,
        position INTEGER DEFAULT 0,
        width INTEGER DEFAULT 6,
        height INTEGER DEFAULT 3,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    # Bid templates (reuse existing if present)
    c.execute('''CREATE TABLE IF NOT EXISTS pm_bid_templates (
        id SERIAL PRIMARY KEY,
        name TEXT NOT NULL UNIQUE,
        description TEXT,
        items JSONB,
        default_markup NUMERIC(6,2) DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    # Risks
    c.execute('''CREATE TABLE IF NOT EXISTS pm_risks (
        id SERIAL PRIMARY KEY,
        job_id INTEGER,
        title TEXT NOT NULL,
        probability INTEGER CHECK (probability BETWEEN 1 AND 5),
        impact INTEGER CHECK (impact BETWEEN 1 AND 5),
        mitigation TEXT,
        owner_id INTEGER,
        status TEXT DEFAULT 'open',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE SET NULL,
        FOREIGN KEY (owner_id) REFERENCES users(id) ON DELETE SET NULL
    )''')
    # Time logs
    c.execute('''CREATE TABLE IF NOT EXISTS pm_time_logs (
        id SERIAL PRIMARY KEY,
        item_id INTEGER NOT NULL,
        user_id INTEGER,
        hours NUMERIC(10,2) NOT NULL,
        work_date DATE DEFAULT CURRENT_DATE,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (item_id) REFERENCES pm_items(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
    )''')
    # Automations
    c.execute('''CREATE TABLE IF NOT EXISTS pm_automations (
        id SERIAL PRIMARY KEY,
        board_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        trigger JSONB,  -- e.g., {"field":"status","from":"Any","to":"Delayed"}
        actions JSONB,  -- e.g., [ {"type":"notify","role":"manager"}, {"type":"update","field":"timeline","value":"+2d"} ]
        monthly_limit INTEGER DEFAULT 250,
        used_this_month INTEGER DEFAULT 0,
        active BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (board_id) REFERENCES pm_boards(id) ON DELETE CASCADE
    )''')
    # Dependencies (edges between items)
    c.execute('''CREATE TABLE IF NOT EXISTS pm_dependencies (
        id SERIAL PRIMARY KEY,
        board_id INTEGER NOT NULL,
        predecessor_id INTEGER NOT NULL,
        successor_id INTEGER NOT NULL,
        type TEXT DEFAULT 'FS', -- FS, SS, FF, SF
        FOREIGN KEY (board_id) REFERENCES pm_boards(id) ON DELETE CASCADE,
        FOREIGN KEY (predecessor_id) REFERENCES pm_items(id) ON DELETE CASCADE,
        FOREIGN KEY (successor_id) REFERENCES pm_items(id) ON DELETE CASCADE,
        UNIQUE (predecessor_id, successor_id)
    )''')
    # Teams and members
    c.execute('''CREATE TABLE IF NOT EXISTS pm_teams (
        id SERIAL PRIMARY KEY,
        name TEXT NOT NULL UNIQUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS pm_team_members (
        id SERIAL PRIMARY KEY,
        team_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        role TEXT DEFAULT 'member',
        UNIQUE(team_id, user_id),
        FOREIGN KEY (team_id) REFERENCES pm_teams(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    # Forms
    c.execute('''CREATE TABLE IF NOT EXISTS pm_forms (
        id SERIAL PRIMARY KEY,
        board_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        definition JSONB, -- fields mapping to columns
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (board_id) REFERENCES pm_boards(id) ON DELETE CASCADE
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS pm_form_submissions (
        id SERIAL PRIMARY KEY,
        form_id INTEGER NOT NULL,
        data JSONB,
        created_item_id INTEGER,
        submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (form_id) REFERENCES pm_forms(id) ON DELETE CASCADE,
        FOREIGN KEY (created_item_id) REFERENCES pm_items(id) ON DELETE SET NULL
    )''')
    # User capacity (weekly hours)
    c.execute('''CREATE TABLE IF NOT EXISTS pm_user_capacity (
        id SERIAL PRIMARY KEY,
        user_id INTEGER NOT NULL UNIQUE,
        weekly_capacity_hours NUMERIC(10,2) DEFAULT 40,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    conn.commit()
    release_connection(conn)

def create_maintenance_extra_tables():
    """Create additional maintenance tables for CRM, inspections, estimating, invoicing, messages, time logs, campaigns."""
    conn = get_database_connection()
    try:
        conn.autocommit = True
        c = conn.cursor()

        # CRM: customers and vehicles/assets
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_customers (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            contact_email TEXT,
            contact_phone TEXT,
            company TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_vehicles (
            id SERIAL PRIMARY KEY,
            customer_id INTEGER,
            equipment_id INTEGER,
            vin TEXT,
            plate TEXT,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES maintenance_customers(id) ON DELETE SET NULL,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id) ON DELETE SET NULL
        )''')

        # Inspections (DVI)
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_inspections (
            id SERIAL PRIMARY KEY,
            equipment_id INTEGER NOT NULL,
            inspector_id INTEGER,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id),
            FOREIGN KEY (inspector_id) REFERENCES users(id)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_inspection_media (
            id SERIAL PRIMARY KEY,
            inspection_id INTEGER NOT NULL,
            media_path TEXT NOT NULL,
            media_type TEXT,
            caption TEXT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (inspection_id) REFERENCES maintenance_inspections(id) ON DELETE CASCADE
        )''')

        # DVI checklist templates and items
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_checklist_templates (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            equipment_type TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_checklist_template_items (
            id SERIAL PRIMARY KEY,
            template_id INTEGER NOT NULL,
            label TEXT NOT NULL,
            default_status TEXT DEFAULT 'ok',
            position INTEGER DEFAULT 0,
            FOREIGN KEY (template_id) REFERENCES maintenance_checklist_templates(id) ON DELETE CASCADE
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_inspection_items (
            id SERIAL PRIMARY KEY,
            inspection_id INTEGER NOT NULL,
            label TEXT NOT NULL,
            status TEXT DEFAULT 'ok',
            notes TEXT,
            position INTEGER DEFAULT 0,
            FOREIGN KEY (inspection_id) REFERENCES maintenance_inspections(id) ON DELETE CASCADE
        )''')

        # Estimates and invoices
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_estimates (
            id SERIAL PRIMARY KEY,
            customer_id INTEGER,
            equipment_id INTEGER,
            job_id INTEGER,
            subtotal NUMERIC(12,2) DEFAULT 0,
            tax NUMERIC(12,2) DEFAULT 0,
            total NUMERIC(12,2) DEFAULT 0,
            status TEXT DEFAULT 'draft',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES maintenance_customers(id),
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_estimate_items (
            id SERIAL PRIMARY KEY,
            estimate_id INTEGER NOT NULL,
            item_type TEXT,
            description TEXT,
            quantity NUMERIC(10,2) DEFAULT 1,
            unit_price NUMERIC(12,2) DEFAULT 0,
            total NUMERIC(12,2) DEFAULT 0,
            FOREIGN KEY (estimate_id) REFERENCES maintenance_estimates(id) ON DELETE CASCADE
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_invoices (
            id SERIAL PRIMARY KEY,
            customer_id INTEGER,
            equipment_id INTEGER,
            estimate_id INTEGER,
            subtotal NUMERIC(12,2) DEFAULT 0,
            tax NUMERIC(12,2) DEFAULT 0,
            total NUMERIC(12,2) DEFAULT 0,
            status TEXT DEFAULT 'unpaid',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES maintenance_customers(id),
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id),
            FOREIGN KEY (estimate_id) REFERENCES maintenance_estimates(id)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_invoice_items (
            id SERIAL PRIMARY KEY,
            invoice_id INTEGER NOT NULL,
            item_type TEXT,
            description TEXT,
            quantity NUMERIC(10,2) DEFAULT 1,
            unit_price NUMERIC(12,2) DEFAULT 0,
            total NUMERIC(12,2) DEFAULT 0,
            FOREIGN KEY (invoice_id) REFERENCES maintenance_invoices(id) ON DELETE CASCADE
        )''')

        # Communications and time logs
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_messages (
            id SERIAL PRIMARY KEY,
            job_id INTEGER,
            sender_id INTEGER,
            message TEXT,
            channel TEXT DEFAULT 'internal',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (job_id) REFERENCES equipment_maintenance_jobs(id) ON DELETE CASCADE,
            FOREIGN KEY (sender_id) REFERENCES users(id) ON DELETE SET NULL
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_time_logs (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL,
            user_id INTEGER,
            hours NUMERIC(10,2) NOT NULL,
            work_date DATE DEFAULT CURRENT_DATE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (job_id) REFERENCES equipment_maintenance_jobs(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
        )''')

        # Campaigns (marketing)
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_campaigns (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            channel TEXT DEFAULT 'email',
            content TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_campaign_messages (
            id SERIAL PRIMARY KEY,
            campaign_id INTEGER NOT NULL,
            recipient TEXT NOT NULL,
            status TEXT DEFAULT 'queued',
            sent_at TIMESTAMP,
            FOREIGN KEY (campaign_id) REFERENCES maintenance_campaigns(id) ON DELETE CASCADE
        )''')

        # =====================================================================
        # ENHANCED MAINTENANCE TABLES (2025 Upgrade)
        # =====================================================================

        # Enhanced checklist templates with more fields
        c.execute('''ALTER TABLE maintenance_checklist_templates
            ADD COLUMN IF NOT EXISTS description TEXT,
            ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE''')

        # Enhanced checklist template items
        c.execute('''ALTER TABLE maintenance_checklist_template_items
            ADD COLUMN IF NOT EXISTS category TEXT,
            ADD COLUMN IF NOT EXISTS requires_photo BOOLEAN DEFAULT FALSE''')

        # Enhanced inspections table
        c.execute('''ALTER TABLE maintenance_inspections
            ADD COLUMN IF NOT EXISTS template_id INTEGER,
            ADD COLUMN IF NOT EXISTS job_id INTEGER,
            ADD COLUMN IF NOT EXISTS inspection_type TEXT DEFAULT 'routine',
            ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'in_progress',
            ADD COLUMN IF NOT EXISTS completed_at TIMESTAMP,
            ADD COLUMN IF NOT EXISTS signature_data TEXT,
            ADD COLUMN IF NOT EXISTS signature_name TEXT''')

        # Enhanced inspection items
        c.execute('''ALTER TABLE maintenance_inspection_items
            ADD COLUMN IF NOT EXISTS category TEXT,
            ADD COLUMN IF NOT EXISTS requires_photo BOOLEAN DEFAULT FALSE''')

        # Enhanced inspection media
        c.execute('''ALTER TABLE maintenance_inspection_media
            ADD COLUMN IF NOT EXISTS item_id INTEGER''')

        # Work Order Templates
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_work_order_templates (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            job_type TEXT NOT NULL,
            equipment_type TEXT,
            description TEXT,
            estimated_hours NUMERIC(10,2),
            safety_notes TEXT,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_work_order_template_steps (
            id SERIAL PRIMARY KEY,
            template_id INTEGER NOT NULL,
            step_number INTEGER NOT NULL,
            instruction TEXT NOT NULL,
            estimated_minutes INTEGER,
            requires_signoff BOOLEAN DEFAULT FALSE,
            FOREIGN KEY (template_id) REFERENCES maintenance_work_order_templates(id) ON DELETE CASCADE
        )''')

        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_work_order_template_parts (
            id SERIAL PRIMARY KEY,
            template_id INTEGER NOT NULL,
            consumable_id INTEGER NOT NULL,
            quantity INTEGER DEFAULT 1,
            notes TEXT,
            FOREIGN KEY (template_id) REFERENCES maintenance_work_order_templates(id) ON DELETE CASCADE
        )''')

        # Job Steps (procedure tracking)
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_job_steps (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL,
            step_number INTEGER NOT NULL,
            instruction TEXT NOT NULL,
            estimated_minutes INTEGER,
            requires_signoff BOOLEAN DEFAULT FALSE,
            status TEXT DEFAULT 'pending',
            completed_at TIMESTAMP,
            completed_by INTEGER,
            signature_data TEXT,
            notes TEXT,
            FOREIGN KEY (job_id) REFERENCES equipment_maintenance_jobs(id) ON DELETE CASCADE
        )''')

        # Job Signatures (e-signatures for compliance)
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_job_signatures (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL,
            signature_type TEXT NOT NULL,
            signature_data TEXT NOT NULL,
            signer_name TEXT NOT NULL,
            signer_id INTEGER,
            signed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY (job_id) REFERENCES equipment_maintenance_jobs(id) ON DELETE CASCADE
        )''')

        # Enhance equipment_maintenance_jobs table
        c.execute('''ALTER TABLE equipment_maintenance_jobs
            ADD COLUMN IF NOT EXISTS template_id INTEGER,
            ADD COLUMN IF NOT EXISTS estimated_hours NUMERIC(10,2),
            ADD COLUMN IF NOT EXISTS acknowledged_at TIMESTAMP,
            ADD COLUMN IF NOT EXISTS resolved_at TIMESTAMP,
            ADD COLUMN IF NOT EXISTS vendor_id INTEGER,
            ADD COLUMN IF NOT EXISTS labor_cost NUMERIC(12,2),
            ADD COLUMN IF NOT EXISTS parts_cost NUMERIC(12,2),
            ADD COLUMN IF NOT EXISTS total_cost NUMERIC(12,2)''')

        # Enhance maintenance_attachments table
        c.execute('''ALTER TABLE maintenance_attachments
            ADD COLUMN IF NOT EXISTS file_type TEXT,
            ADD COLUMN IF NOT EXISTS uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP''')

        # Vendors and Contractors
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_vendors (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            vendor_type TEXT DEFAULT 'contractor',
            contact_name TEXT,
            contact_email TEXT,
            contact_phone TEXT,
            address TEXT,
            specialties TEXT,
            rating NUMERIC(3,2),
            notes TEXT,
            is_active BOOLEAN DEFAULT TRUE,
            insurance_expiry DATE,
            contract_expiry DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        c.execute('''CREATE TABLE IF NOT EXISTS maintenance_vendor_jobs (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL,
            vendor_id INTEGER NOT NULL,
            estimated_cost NUMERIC(12,2),
            cost NUMERIC(12,2),
            status TEXT DEFAULT 'assigned',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            FOREIGN KEY (job_id) REFERENCES equipment_maintenance_jobs(id) ON DELETE CASCADE,
            FOREIGN KEY (vendor_id) REFERENCES maintenance_vendors(id) ON DELETE CASCADE
        )''')

        # Equipment Certifications
        c.execute('''CREATE TABLE IF NOT EXISTS equipment_certifications (
            id SERIAL PRIMARY KEY,
            equipment_id INTEGER NOT NULL,
            certification_type TEXT NOT NULL,
            certification_number TEXT,
            issued_date DATE,
            expiry_date DATE,
            status TEXT DEFAULT 'valid',
            notes TEXT,
            document_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id) ON DELETE CASCADE
        )''')

        # Equipment Warranties
        c.execute('''CREATE TABLE IF NOT EXISTS equipment_warranties (
            id SERIAL PRIMARY KEY,
            equipment_id INTEGER NOT NULL,
            warranty_type TEXT NOT NULL,
            provider TEXT NOT NULL,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            coverage_details TEXT,
            notes TEXT,
            document_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id) ON DELETE CASCADE
        )''')

        # Citizen/311 Requests
        c.execute('''CREATE TABLE IF NOT EXISTS citizen_requests (
            id SERIAL PRIMARY KEY,
            request_type TEXT NOT NULL,
            description TEXT NOT NULL,
            location TEXT,
            latitude NUMERIC(10,7),
            longitude NUMERIC(10,7),
            reporter_name TEXT,
            reporter_email TEXT,
            reporter_phone TEXT,
            photo_path TEXT,
            status TEXT DEFAULT 'new',
            job_id INTEGER,
            resolution_notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            resolved_at TIMESTAMP,
            FOREIGN KEY (job_id) REFERENCES equipment_maintenance_jobs(id) ON DELETE SET NULL
        )''')

        # Telematics Data
        c.execute('''CREATE TABLE IF NOT EXISTS equipment_telematics (
            id SERIAL PRIMARY KEY,
            equipment_id INTEGER NOT NULL,
            data_type TEXT NOT NULL,
            value NUMERIC(15,4) NOT NULL,
            unit TEXT,
            source TEXT,
            raw_data JSONB,
            recorded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id) ON DELETE CASCADE
        )''')

        # Telematics Alert Rules
        c.execute('''CREATE TABLE IF NOT EXISTS telematics_alert_rules (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            equipment_type TEXT,
            data_type TEXT NOT NULL,
            alert_type TEXT DEFAULT 'warning',
            threshold_min NUMERIC(15,4),
            threshold_max NUMERIC(15,4),
            alert_message TEXT,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        # Telematics Alerts
        c.execute('''CREATE TABLE IF NOT EXISTS telematics_alerts (
            id SERIAL PRIMARY KEY,
            equipment_id INTEGER NOT NULL,
            rule_id INTEGER,
            data_type TEXT NOT NULL,
            value NUMERIC(15,4),
            alert_message TEXT,
            status TEXT DEFAULT 'new',
            acknowledged_by INTEGER,
            acknowledged_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id) ON DELETE CASCADE
        )''')

        # Geofences
        c.execute('''CREATE TABLE IF NOT EXISTS geofences (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            coordinates JSONB NOT NULL,
            alert_on_exit BOOLEAN DEFAULT TRUE,
            alert_on_enter BOOLEAN DEFAULT FALSE,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        c.execute('''CREATE TABLE IF NOT EXISTS geofence_equipment (
            id SERIAL PRIMARY KEY,
            geofence_id INTEGER NOT NULL,
            equipment_id INTEGER NOT NULL,
            FOREIGN KEY (geofence_id) REFERENCES geofences(id) ON DELETE CASCADE,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id) ON DELETE CASCADE
        )''')

        # Create indexes for performance
        c.execute('''CREATE INDEX IF NOT EXISTS idx_telematics_equipment_type
            ON equipment_telematics(equipment_id, data_type, recorded_at DESC)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_inspections_equipment
            ON maintenance_inspections(equipment_id, created_at DESC)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_jobs_status
            ON equipment_maintenance_jobs(status, due_date)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_citizen_requests_status
            ON citizen_requests(status, created_at DESC)''')

        # ======================================================================
        # Parts Kits and Cross-Reference Tables for Enhanced Inventory
        # ======================================================================

        # Parts Kits - bundled groups of items for common maintenance tasks
        c.execute('''CREATE TABLE IF NOT EXISTS parts_kits (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            kit_type TEXT DEFAULT 'maintenance',
            equipment_type TEXT,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by INTEGER
        )''')

        # Parts Kit Items - items included in each kit
        c.execute('''CREATE TABLE IF NOT EXISTS parts_kit_items (
            id SERIAL PRIMARY KEY,
            kit_id INTEGER NOT NULL,
            consumable_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 1,
            is_optional BOOLEAN DEFAULT FALSE,
            notes TEXT,
            FOREIGN KEY (kit_id) REFERENCES parts_kits(id) ON DELETE CASCADE,
            FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
            UNIQUE (kit_id, consumable_id)
        )''')

        # Parts Cross-Reference - equivalent parts from different manufacturers
        c.execute('''CREATE TABLE IF NOT EXISTS parts_cross_reference (
            id SERIAL PRIMARY KEY,
            primary_consumable_id INTEGER NOT NULL,
            equivalent_consumable_id INTEGER NOT NULL,
            match_type TEXT DEFAULT 'equivalent',
            notes TEXT,
            verified BOOLEAN DEFAULT FALSE,
            verified_by INTEGER,
            verified_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (primary_consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
            FOREIGN KEY (equivalent_consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
            UNIQUE (primary_consumable_id, equivalent_consumable_id)
        )''')

        # Job Parts Reserved - track parts reserved for specific jobs
        c.execute('''CREATE TABLE IF NOT EXISTS job_parts_reserved (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL,
            consumable_id INTEGER NOT NULL,
            quantity_reserved INTEGER NOT NULL,
            reserved_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'reserved',
            used_at TIMESTAMP,
            notes TEXT,
            FOREIGN KEY (job_id) REFERENCES equipment_maintenance_jobs(id) ON DELETE CASCADE,
            FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE,
            UNIQUE (job_id, consumable_id)
        )''')

        # Add enhanced stock level columns to consumables if not exists
        c.execute('''ALTER TABLE consumables
            ADD COLUMN IF NOT EXISTS safety_stock INTEGER,
            ADD COLUMN IF NOT EXISTS reorder_quantity INTEGER,
            ADD COLUMN IF NOT EXISTS lead_time_days INTEGER,
            ADD COLUMN IF NOT EXISTS abc_classification TEXT,
            ADD COLUMN IF NOT EXISTS last_count_date TIMESTAMP,
            ADD COLUMN IF NOT EXISTS last_order_date TIMESTAMP''')

        # Inventory Count Records - for periodic inventory counts
        c.execute('''CREATE TABLE IF NOT EXISTS inventory_counts (
            id SERIAL PRIMARY KEY,
            consumable_id INTEGER NOT NULL,
            count_type TEXT DEFAULT 'cycle',
            system_quantity INTEGER NOT NULL,
            counted_quantity INTEGER NOT NULL,
            variance INTEGER GENERATED ALWAYS AS (counted_quantity - system_quantity) STORED,
            counted_by INTEGER,
            counted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            adjustment_made BOOLEAN DEFAULT FALSE,
            FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE
        )''')

        # Create indexes for new tables
        c.execute('''CREATE INDEX IF NOT EXISTS idx_parts_kits_type
            ON parts_kits(kit_type, equipment_type)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_cross_reference_primary
            ON parts_cross_reference(primary_consumable_id)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_job_parts_reserved
            ON job_parts_reserved(job_id, status)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_consumables_stock_level
            ON consumables(quantity, minimum_stock_level, reorder_threshold) WHERE active = TRUE''')

        # ======================================================================
        # Telematics Provider Configuration Tables
        # ======================================================================

        # Telematics Providers - GPS/Fleet tracking provider configurations
        c.execute('''CREATE TABLE IF NOT EXISTS telematics_providers (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            provider_type TEXT NOT NULL,
            config JSONB,
            is_active BOOLEAN DEFAULT TRUE,
            last_sync TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        # Equipment Telematics Links - map local equipment to provider vehicle IDs
        c.execute('''CREATE TABLE IF NOT EXISTS equipment_telematics_links (
            id SERIAL PRIMARY KEY,
            equipment_id INTEGER NOT NULL,
            provider_id INTEGER NOT NULL,
            external_id TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (equipment_id) REFERENCES equipment_instances(id) ON DELETE CASCADE,
            FOREIGN KEY (provider_id) REFERENCES telematics_providers(id) ON DELETE CASCADE,
            UNIQUE (equipment_id, provider_id)
        )''')

        # Index for efficient lookups
        c.execute('''CREATE INDEX IF NOT EXISTS idx_telematics_links_provider
            ON equipment_telematics_links(provider_id, external_id)''')

        conn.commit()

    finally:
        try:
            release_connection(conn)
        except Exception:
            pass
# Update the main init_db() function to include these
def init_db():
    conn = None
    try:
        conn = get_connection()
        c = conn.cursor()

        # Core system tables must exist first (jobs, equipment, etc.)
        create_basic_tables()

        # Add the missing inventory tables
        create_inventory_alerts_table()
        create_job_consumables_table()
        create_additional_inventory_tables()
        # Ensure purchasing + reorder workflow tables exist
        create_purchasing_integration_tables()
        create_reorder_tables()
        create_default_templates()
        # Project Management core tables
        create_pm_tables()
        # Maintenance extra tables
        create_maintenance_extra_tables()
        # Tables required by automation_service background jobs
        create_automation_email_tables()
        # User roles and permissions tables
        create_user_roles_tables()
        # Billing tables
        create_billing_tables()

        # Import Excel data
        import_service_units_from_excel()

        # Dedupe users after table creation
        dedupe_users()

        # ... rest of existing init_db code ...

        # Seed core sample data (admin user, etc.) idempotently
        try:
            insert_sample_data()
        except Exception as _seed_err:
            logging.warning(f"Seeding sample data skipped/failed: {_seed_err}")

        conn.commit()
        logging.info("Database initialization completed successfully")
    except Exception as e:
        logging.error(f"Database initialization failed: {str(e)}")
        raise
    finally:
        if conn is not None:
            try:
                release_connection(conn)
            except Exception:
                pass

def create_automation_email_tables():
    """Create email templates, automation rules, and email logs tables."""
    conn = None
    try:
        conn = get_connection()
        conn.autocommit = True
        c = conn.cursor()

        c.execute('''
            CREATE TABLE IF NOT EXISTS email_templates (
                id SERIAL PRIMARY KEY,
                subject TEXT NOT NULL,
                body TEXT NOT NULL,
                outlook_enabled BOOLEAN DEFAULT FALSE,
                cc TEXT,
                bcc TEXT,
                is_html BOOLEAN DEFAULT FALSE,
                last_used TIMESTAMP NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS automation_rules (
                id SERIAL PRIMARY KEY,
                template_id INTEGER REFERENCES email_templates(id) ON DELETE SET NULL,
                trigger_type TEXT NOT NULL,
                trigger_value TEXT,
                recipient_type TEXT,
                recipient_value TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                active BOOLEAN DEFAULT TRUE
            )
        ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS email_logs (
                id SERIAL PRIMARY KEY,
                template_id INTEGER REFERENCES email_templates(id) ON DELETE SET NULL,
                job_id INTEGER,
                recipient TEXT,
                subject TEXT,
                sent_timestamp TIMESTAMP,
                status TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
    except Exception as e:
        logging.error(f"Error creating automation/email tables: {str(e)}")
    finally:
        try:
            if conn:
                release_connection(conn)
        except Exception:
            pass

def create_user_roles_tables():
    """Create tables for roles, permissions, and user-role assignments."""
    conn = None
    try:
        conn = get_connection()
        conn.autocommit = True
        c = conn.cursor()

        # Roles table
        c.execute('''
            CREATE TABLE IF NOT EXISTS roles (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Permissions table (if needed; or store as JSON in roles)
        c.execute('''
            CREATE TABLE IF NOT EXISTS permissions (
                id SERIAL PRIMARY KEY,
                role_id INTEGER REFERENCES roles(id) ON DELETE CASCADE,
                permission TEXT NOT NULL,
                UNIQUE(role_id, permission)
            )
        ''')

        # User-roles junction
        c.execute('''
            CREATE TABLE IF NOT EXISTS user_roles (
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                role_id INTEGER REFERENCES roles(id) ON DELETE CASCADE,
                PRIMARY KEY (user_id, role_id)
            )
        ''')

        # Preset roles
        preset_roles = [
            ('admin', 'Full access'),
            ('manager', 'Manage users and projects'),
            ('worker', 'Basic access'),
            ('maintenance', 'Maintenance tasks')
        ]
        for name, desc in preset_roles:
            c.execute('''
                INSERT INTO roles (name, description)
                VALUES (%s, %s)
                ON CONFLICT (name) DO NOTHING
            ''', (name, desc))

    except Exception as e:
        logging.error(f"Error creating user roles tables: {str(e)}")
    finally:
        try:
            if conn:
                release_connection(conn)
        except Exception:
            pass

def dedupe_users():
    """Remove duplicate users, keeping the one with the lowest ID."""
    conn = None
    try:
        conn = get_connection()
        conn.autocommit = True
        c = conn.cursor()
        c.execute('''
            DELETE FROM users a
            USING users b
            WHERE a.id > b.id
            AND a.username = b.username
            AND a.email = b.email
        ''')
        logging.info(f"Deduplicated {c.rowcount} user records")
    except Exception as e:
        logging.error(f"Error deduping users: {str(e)}")
    finally:
        try:
            if conn:
                release_connection(conn)
        except Exception:
            pass

def fix_database_schema_inconsistencies():
    """Fix any remaining database schema issues"""
    conn = get_connection()
    c = conn.cursor()
    
    try:
        # Ensure all required columns exist in consumables table
        required_columns = {
            'id': 'SERIAL PRIMARY KEY',
            'name': 'TEXT NOT NULL',
            'category': 'TEXT',
            'base_name': 'TEXT',
            'specifications': 'JSONB',
            'location': 'TEXT',
            'bin_location': 'TEXT',
            'quantity': 'INTEGER DEFAULT 0',
            'supplier': 'TEXT',
            'manufacturer': 'TEXT',
            'part_number': 'TEXT',
            'supplier_sku': 'TEXT',
            'internal_sku': 'TEXT',
            'serial_numbers': 'TEXT',  # Store as JSON string
            'unit': 'TEXT DEFAULT \'each\'',
            'reorder_threshold': 'INTEGER',
            'minimum_stock_level': 'INTEGER',
            'maximum_stock_level': 'INTEGER',
            'cost_per_unit': 'DECIMAL(10,4)',
            'weight_per_unit': 'DECIMAL(10,4)',
            'dimensions': 'JSONB',
            'barcode': 'TEXT',
            'image_path': 'TEXT',
            'image_filename': 'TEXT',
            'hazmat': 'BOOLEAN DEFAULT FALSE',
            'requires_certification': 'BOOLEAN DEFAULT FALSE',
            'product_url': 'TEXT',
            'notes': 'TEXT',
            'normalized_name': 'TEXT',
            'active': 'BOOLEAN DEFAULT TRUE',
            'created_at': 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP',
            'updated_at': 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP'
        }
        
        # Check existing columns
        c.execute("""
            SELECT column_name, data_type 
            FROM information_schema.columns 
            WHERE table_name = 'consumables'
        """)
        existing_columns = {row[0]: row[1] for row in c.fetchall()}
        
        # Add missing columns
        for col_name, col_definition in required_columns.items():
            if col_name not in existing_columns:
                try:
                    # Skip PRIMARY KEY constraint for existing tables
                    if 'PRIMARY KEY' in col_definition and existing_columns:
                        col_definition = col_definition.replace(' PRIMARY KEY', '')
                    c.execute(f"ALTER TABLE consumables ADD COLUMN {col_name} {col_definition}")
                    logging.info(f"Added column {col_name} to consumables table")
                except Exception as e:
                    logging.warning(f"Could not add column {col_name}: {str(e)}")
        
        # Create missing indexes
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_consumables_category ON consumables(category)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_base_name ON consumables(base_name)", 
            "CREATE INDEX IF NOT EXISTS idx_consumables_supplier ON consumables(supplier)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_internal_sku ON consumables(internal_sku)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_supplier_sku ON consumables(supplier_sku)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_active ON consumables(active)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_quantity ON consumables(quantity)",
            "CREATE INDEX IF NOT EXISTS idx_consumables_created_at ON consumables(created_at)"
        ]
        
        for index_sql in indexes:
            try:
                c.execute(index_sql)
            except Exception as e:
                logging.warning(f"Could not create index: {str(e)}")
        
        # Ensure jobs table has required columns for ML predictions
        c.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'jobs'
        """)
        job_columns = [row[0] for row in c.fetchall()]
        
        job_required_columns = [
            ('job_type', 'TEXT'),
            ('square_footage', 'INTEGER'),
            ('crew_size', 'INTEGER'),
            ('duration_days', 'INTEGER')
        ]
        
        for col_name, col_type in job_required_columns:
            if col_name not in job_columns:
                try:
                    c.execute(f"ALTER TABLE jobs ADD COLUMN {col_name} {col_type}")
                    logging.info(f"Added column {col_name} to jobs table")
                except Exception as e:
                    logging.warning(f"Could not add column {col_name} to jobs: {str(e)}")
        
        conn.commit()
        logging.info("Database schema fixes completed successfully")
        
    except Exception as e:
        conn.rollback()
        logging.error(f"Database schema fixes failed: {str(e)}")
        raise
    finally:
        release_connection(conn)
        
def insert_sample_data():
    """Insert sample data for testing"""
    conn = get_database_connection()
    c = conn.cursor()
    
    try:
        logger.info("🔧 Inserting sample data...")
        
        # Sample users
        sample_users = [
            ("admin", "Admin User", "admin@potelco.com", "System Administrator", "admin", 
             generate_password_hash("potelco123"), '["forklift", "commercial_driver"]'),
            ("manager1", "Jane Manager", "manager1@potelco.com", "Site Manager", "manager", 
             generate_password_hash("mgr2025"), '["forklift"]'),
            ("user1", "John Worker", "user1@potelco.com", "Operator", "user", 
             generate_password_hash("worker123"), '["boom_operator"]')
        ]
        
        for user in sample_users:
            c.execute("""
                INSERT INTO users (username, full_name, email, job_title, role, password, certifications) 
                VALUES (%s, %s, %s, %s, %s, %s, %s) 
                ON CONFLICT (username) DO NOTHING
            """, user)
        
        # Sample towns
        towns = [
            ("Seattle", "permits@seattle.gov", "City Permits", "206-555-1000", "600 4th Ave, Seattle, WA"),
            ("Tacoma", "permits@tacoma.gov", "City Permits", "253-555-2000", "747 Market St, Tacoma, WA"),
        ]
        for name, email, contact_name, phone, address in towns:
            c.execute(
                """
                INSERT INTO city_contacts (city_name, email, contact_name, phone, address)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (city_name) DO UPDATE SET email = EXCLUDED.email, contact_name = EXCLUDED.contact_name
                """,
                (name, email, contact_name, phone, address),
            )

        # Town contacts
        # Find IDs
        c.execute("SELECT id, city_name FROM city_contacts")
        town_rows = c.fetchall()
        name_to_id = {r[1]: r[0] for r in town_rows}
        seattle_id = name_to_id.get("Seattle")
        tacoma_id = name_to_id.get("Tacoma")
        if seattle_id:
            c.execute(
                """
                INSERT INTO town_contacts (town_id, role, name, email, phone)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT DO NOTHING
                """,
                (seattle_id, "digging", "John DigSafe", "dig@seattle.gov", "206-555-1010"),
            )
            c.execute(
                """
                INSERT INTO town_contacts (town_id, role, name, email)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT DO NOTHING
                """,
                (seattle_id, "road_closures", "Jane ROW", "row@seattle.gov"),
            )

        # Sample equipment
        sample_equipment = [
            ("Boom Truck", "BT001", "Ford", "F-550", "1FT8W3BTXJEB12345", 0, "diesel", 19500, False, "boom_operator", "available"),
            ("Boom Truck", "BT002", "Ford", "F-550", "1FT8W3BTXJEB12346", 0, "diesel", 19500, False, "boom_operator", "available"),
            ("Crane", "CR001", "Terex", "T340-1XL", "123456789", 0, "diesel", 66000, False, "crane_certified", "available")
        ]
        
        for eq in sample_equipment:
            c.execute("""
                INSERT INTO equipment_instances 
                (equipment_type, unique_id, brand, model, serial_number, hours, fuel_type, 
                 gross_weight, requires_operator, required_certification, status) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) 
                ON CONFLICT (unique_id) DO NOTHING
            """, eq)
        
        # Sample consumables
        consumables = [
            ("5/8\" x 6\" Galv Bolt", "Warehouse A", 120, "Fastenal", None, "each", 50),
            ("#2 CU THHN 500ft Reel", "Yard B", 4, "Graybar", None, "reel", 2),
            ("Safety Gloves Class 2", "Warehouse A", 40, "PPE Co", None, "pairs", 20),
        ]
        for name, loc, qty, supplier, serials, unit, reorder in consumables:
            c.execute(
                """
                INSERT INTO consumables (name, location, quantity, supplier, serial_numbers, unit, reorder_threshold)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT DO NOTHING
                """,
                (name, loc, qty, supplier, serials, unit, reorder),
            )

        # Sample job with town linkage
        seattle_town = seattle_id if seattle_id else None
        c.execute(
            """
            INSERT INTO jobs (name, description, job_type, estimated_cost, location, town_id, status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT DO NOTHING
            """,
            ("Pole Installation Job", "Install utility pole in Seattle", "installation", 5000.0, "Seattle", seattle_town, "pending", datetime.now()),
        )

        # Sample schedules
        c.execute(
            """
            INSERT INTO schedules (title, start_date, end_date, description, user_id, location, resource_name, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT DO NOTHING
            """,
            ("Seattle Job", datetime.now().date(), datetime.now().date(), "Set pole", 1, "Seattle", 1, "scheduled"),
        )

        # Sample maintenance job and logs
        c.execute(
            """
            INSERT INTO equipment_maintenance_jobs (equipment_id, job_type, priority, status, assigned_to, due_date, notes, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (1, 'breakdown', 'high', 'new', 2, datetime.now().date(), 'Hydraulic leak reported', 2),
        )
        maint_job_id = c.fetchone()[0]
        c.execute(
            """
            INSERT INTO equipment_maintenance (equipment_id, action, details, status, category, job_id)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (1, 'Inspect', 'Initial inspection scheduled', 'in_progress', 'breakdown', maint_job_id),
        )
        
        conn.commit()
        logger.info("✅ Sample data inserted successfully")
        
    except Exception as e:
        conn.rollback()
        logger.error(f"❌ Error inserting sample data: {e}")
        # Don't raise here - sample data is optional
    finally:
        release_connection(conn)        

# Ensure upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def create_billing_tables():
    """Create tables for service units and invoices."""
    conn = None
    try:
        conn = get_connection()
        conn.autocommit = True
        c = conn.cursor()

        # Service Units from Excel
        c.execute('''
            CREATE TABLE IF NOT EXISTS service_units (
                id SERIAL PRIMARY KEY,
                source_file TEXT,
                service_unit TEXT NOT NULL,
                core_code TEXT,
                service_master TEXT,
                unit_of_measure TEXT,
                definition TEXT,
                clarification TEXT,
                sap_service_master_ TEXT,
                task_code_description TEXT,
                unnamed_2 TEXT,
                description TEXT,
                uom TEXT,
                explanation TEXT,
                om_contract_service_units TEXT,
                core_service_master TEXT,
                non_core_code TEXT,
                non_core_service_master TEXT,
                work_type TEXT,
                definitions TEXT,
                category TEXT,  -- e.g., 'NCC', 'System', etc.
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Invoices
        c.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                job_id INTEGER REFERENCES jobs(id) ON DELETE SET NULL,
                invoice_number TEXT UNIQUE,
                title TEXT,
                customer_id INTEGER REFERENCES maintenance_customers(id),
                status TEXT DEFAULT 'draft',
                subtotal NUMERIC(12,2) DEFAULT 0,
                tax NUMERIC(12,2) DEFAULT 0,
                total NUMERIC(12,2) DEFAULT 0,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                sent_at TIMESTAMP,
                paid_at TIMESTAMP
            )
        ''')

        # Invoice line items (linked to service units)
        c.execute('''
            CREATE TABLE IF NOT EXISTS invoice_lines (
                id SERIAL PRIMARY KEY,
                invoice_id INTEGER REFERENCES invoices(id) ON DELETE CASCADE,
                service_unit_id INTEGER REFERENCES service_units(id),
                description TEXT,
                quantity NUMERIC(10,2) DEFAULT 1,
                unit_price NUMERIC(12,2) DEFAULT 0,
                total NUMERIC(12,2) DEFAULT 0,
                position INTEGER DEFAULT 0
            )
        ''')

    except Exception as e:
        logging.error(f"Error creating billing tables: {str(e)}")
    finally:
        try:
            if conn:
                release_connection(conn)
        except Exception:
            pass

def import_service_units_from_excel():
    """Import data from Excel into service_units table."""
    wb = openpyxl.load_workbook('4.2_ 2025 Clarification Docs_NO PRICING.xlsx', read_only=True)
    conn = get_connection()
    c = conn.cursor()
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]
            if not headers:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                c.execute('''
                    INSERT INTO service_units (service_unit, core_code, service_master, unit_of_measure, definition, clarification, category)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                ''', (data.get('Service Unit'), data.get('Core Code'), data.get('Service Master'),
                      data.get('Unit of Measure'), data.get('Definition'), data.get('Clarification'), sheet_name))
        conn.commit()
    except Exception as e:
        conn.rollback()
        logging.error(f"Excel import error: {str(e)}")
    finally:
        release_connection(conn)
        wb.close()

if __name__ == "__main__":
    success = init_db()
    sys.exit(0 if success else 1)
